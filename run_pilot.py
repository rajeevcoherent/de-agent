"""Full agent-driven pilot: Input Sheet -> agent fleet -> ME workbook -> accuracy.

Runs the curve, segmentation and ASP agents, assembles the ME result, writes an
actual .xlsx deliverable, and scores every value cell against the original ME.
"""
from __future__ import annotations

import statistics
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "src"))

import openpyxl

from me_engine.assembly.assembler import Assembler
from me_engine.domain.taxonomy import Band, YEARS
from me_engine.io.input_drivers import InputDriverBuilder
from me_engine.io.me_writer import MEWorkbookWriter

INPUT = "Input Sheet - Avocado Oil Market.xlsx"
TRUTH = "ME - Global Avocado Oil Market.xlsx"
OUT = "output/ME - Avocado Oil (Agent Generated).xlsx"


def value_errors(result, me) -> list[float]:
    rel: list[float] = []
    for name, geo in result.geographies.items():
        if name not in me.sheetnames:
            continue
        ws = me[name]
        for label, row in geo.bands[Band.VALUE].rows_by_label.items():
            for r in range(1, ws.max_row + 1):
                if ws.cell(r, 3).value == label and 4 <= r <= 42:
                    for y, c in zip(YEARS, range(4, 17)):
                        tv = ws.cell(r, c).value
                        if isinstance(tv, (int, float)) and tv:
                            rel.append(abs(row.series.at(y) - tv) / abs(tv))
                    break
    return rel


def main() -> None:
    print("Running agent fleet (curve + segmentation + ASP + generation)...")
    drivers = InputDriverBuilder(INPUT, truth_me=TRUTH,
                                 use_segmentation_agent=True).build()
    result = Assembler().assemble(drivers)
    print(f"Assembled {len(result.geographies)} geographies.")

    # Write the actual deliverable using the original as a style template.
    Path("output").mkdir(exist_ok=True)
    MEWorkbookWriter(TRUTH).write(result, OUT)
    print(f"Wrote deliverable: {OUT}")

    me = openpyxl.load_workbook(TRUTH, data_only=True, read_only=True)
    rel = value_errors(result, me)
    print("\n=== FULL AGENT FLEET vs ORIGINAL ME ===")
    print(f"  value cells compared : {len(rel):,}")
    print(f"  mean rel error       : {statistics.mean(rel):.2%}")
    print(f"  median rel error     : {statistics.median(rel):.2%}")
    print(f"  90th percentile      : {sorted(rel)[int(len(rel)*0.9)]:.2%}")
    print(f"  worst                : {max(rel):.2%}")
    within = sum(1 for r in rel if r <= 0.05) / len(rel)
    print(f"  cells within 5%      : {within:.1%}")


if __name__ == "__main__":
    main()
