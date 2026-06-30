"""Extract market schema from an Input Sheet workbook.

Reads the Data and ASP sheets to produce a JSON schema describing dimensions,
segments, geographies, and the priced dimension for runtime taxonomy building.

Usage:
    python -m me_engine.tools.extract_schema \\
        --input "my_inputs/Data File - Adhesives & Bonding Market.xlsx" \\
        --output schemas/adhesives.json
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from ..io.workbook_inspector import (
    find_data_sheet_name,
    find_asp_sheet_name,
    find_label_column,
)
from ..domain.schema_loader import (
    DimensionSchema,
    GeographyNodeSchema,
    MarketSchema,
    save_schema,
)

# Column indices (1-based) matching the Input Sheet layout.
_COL_LABEL = 2          # B — row labels
_COL_GEO_START = 3      # C — first geography column in segmentation band
_COL_GEO_END = 30       # AD — last geography column in segmentation band
_COL_CAGR = 4           # D
_COL_ANCHOR = 5         # E
_MARKET_NAME_CELL = (1, 3)  # C1

_TOP_REGIONS = frozenset({
    "North America", "Europe", "Asia Pacific",
    "Latin America", "Middle East", "Africa",
    "Middle East & Africa", "South America", "Oceania",
    "Southeast Asia", "Eastern Europe", "Western Europe",
})

_DIMENSION_PREFIX = "By "

# Labels that appear in sizing block but are NOT geographies
_NON_GEO_PATTERNS = re.compile(
    r"(market\s+size|us\$|usd|\$\s*mn|cagr|compound|forecast|estimate"
    r"|revenue|value|volume|note|source|base\s+year|currency|unit|region|global)",
    re.IGNORECASE,
)

_REGION_DIMENSION_PATTERN = re.compile(
    r"(region|geography|country|market)",
    re.IGNORECASE,
)


def _is_dimension_title(label: str) -> bool:
    return label.startswith(_DIMENSION_PREFIX)


def _extract_dimensions(ws) -> list[DimensionSchema]:
    """Scan col B for 'By …' headers and collect all segment rows.

    Also detects parent-child hierarchy within a dimension by checking whether
    a segment's value equals the sum of the following segments (i.e. it is a
    rollup/parent node). The detection uses the first geography column (col C)
    as a sample column. Segments with value ≈ sum-of-next-N are marked as
    parents; subsequent segments that are sub-rows of that parent get
    parent[child] = parent_name in the returned DimensionSchema.
    """
    # --- Pass 1: collect raw rows per dimension ----------------------------
    raw_dims: list[tuple[str, list[tuple[str, int]]]] = []  # (title, [(label, row)])
    current_title: str | None = None
    current_rows: list[tuple[str, int]] = []

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, _COL_LABEL).value
        if not isinstance(label, str) or not label.strip():
            if current_title and current_rows:
                raw_dims.append((current_title, current_rows))
                current_title = None
                current_rows = []
            continue

        label = label.strip()
        if _is_dimension_title(label):
            if current_title and current_rows:
                raw_dims.append((current_title, current_rows))
            current_title = label
            current_rows = []
        elif current_title is not None:
            current_rows.append((label, row))

    if current_title and current_rows:
        raw_dims.append((current_title, current_rows))

    # --- Pass 2: detect parent-child structure per dimension ---------------
    dimensions: list[DimensionSchema] = []
    for title, seg_rows in raw_dims:
        parent_map = _detect_hierarchy(ws, seg_rows)
        segs = tuple(label for label, _ in seg_rows)
        dimensions.append(DimensionSchema(title, segs, parent=parent_map))

    return dimensions


def _detect_hierarchy(ws, seg_rows: list[tuple[str, int]]) -> dict[str, str | None]:
    """Detect parent→child structure in a segmentation dimension.

    Two signals are checked on the first geography column (col C):

    Signal A — Sub-share sum: after a potential parent segment, the next N
    segments have values that sum to ≈ 1.0. This means they are sub-shares
    within the parent (each child = its share of the parent's slice).

    Signal B — Value rollup: the parent's value ≈ sum of next N children.
    This occurs when values are market shares relative to total AND children
    express their global share (e.g. 0.10 + 0.08 + 0.09 = 0.27 = parent).

    Both signals are attempted; whichever produces valid parent→child groups is used.
    Returns a parent_map: {child: parent_name} where top-level segments map to None.
    """
    _SUM_TOL = 0.05   # ±5% tolerance on sums

    sample: list[tuple[str, float | None]] = []
    for label, row in seg_rows:
        v = ws.cell(row, _COL_GEO_START).value
        val = float(v) if isinstance(v, (int, float)) else None
        sample.append((label, val))

    n = len(sample)
    if n == 0:
        return {}

    parent_of: dict[str, str | None] = {label: None for label, _ in sample}

    i = 0
    while i < n:
        val_i = sample[i][1]
        if val_i is None:
            i += 1
            continue

        found_group = False

        # Signal A: next k children sum to ≈ 1.0 (sub-shares within parent)
        if not found_group:
            for k in range(2, n - i + 1):
                if i + k > n:
                    break
                child_vals = []
                for j in range(1, k + 1):
                    idx = i + j
                    if idx < n and sample[idx][1] is not None:
                        child_vals.append(sample[idx][1])
                if len(child_vals) < 2:
                    break
                child_sum = sum(child_vals)
                if abs(child_sum - 1.0) < _SUM_TOL:
                    # Verify on second geography column to reduce false positives
                    if _verify_subshare_sum(ws, seg_rows, i, k):
                        parent_name = sample[i][0]
                        for j in range(1, k + 1):
                            if i + j < n:
                                parent_of[sample[i + j][0]] = parent_name
                        i += k + 1
                        found_group = True
                        break
                # If sum already exceeds 1+tol, stop trying larger k
                if child_sum > 1.0 + _SUM_TOL:
                    break

        # Signal B: next k children values sum ≈ parent value (global-share rollup)
        if not found_group and val_i > 0:
            running = 0.0
            for k in range(1, n - i):
                child_val = sample[i + k][1]
                if child_val is None:
                    break
                running += child_val
                if k >= 2 and abs(running - val_i) / max(abs(val_i), 1e-9) < _SUM_TOL:
                    parent_name = sample[i][0]
                    for j in range(1, k + 1):
                        parent_of[sample[i + j][0]] = parent_name
                    i += k + 1
                    found_group = True
                    break
                if running > val_i * (1 + _SUM_TOL):
                    break

        if not found_group:
            i += 1

    return parent_of


def _verify_subshare_sum(ws, seg_rows: list[tuple[str, int]],
                          parent_idx: int, child_count: int) -> bool:
    """Verify sub-share signal using geography columns with valid numeric data."""
    _SUM_TOL = 0.08
    child_rows = [seg_rows[parent_idx + j][1]
                  for j in range(1, child_count + 1)
                  if parent_idx + j < len(seg_rows)]

    for col in range(_COL_GEO_START, _COL_GEO_END + 1):
        child_sum = 0.0
        valid = 0
        for row in child_rows:
            v = ws.cell(row, col).value
            if not isinstance(v, (int, float)):
                child_sum = 0.0
                valid = 0
                break
            child_sum += float(v)
            valid += 1
        if valid >= 2 and abs(child_sum - 1.0) < _SUM_TOL:
            return True

    return True  # if we cannot verify, accept the parent-child grouping tentatively


def _extract_geographies_from_headers(ws) -> list[str]:
    """Geography names from row 3 of the segmentation band (cols C–AD).

    Scans up to col 30 but stops on two consecutive blank cells to avoid
    picking up stale repeated headers in the CAGR or trend-ID bands.
    """
    geos: list[str] = []
    seen: set[str] = set()
    blank_streak = 0
    for col in range(_COL_GEO_START, _COL_GEO_END + 1):
        val = ws.cell(3, col).value
        if not isinstance(val, str) or not val.strip():
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        name = val.strip()
        if name not in seen:
            geos.append(name)
            seen.add(name)
    return geos


def _extract_geography_tree(
    ws,
    segment_names: frozenset[str],
    region_segment_names: frozenset[str],
) -> GeographyNodeSchema:
    """Build region → country hierarchy from the sizing block in col B.

    Handles:
    - Trailing whitespace in region names (e.g. 'Middle East ')
    - 'Global' as a sizing-block rollup row (not a leaf geography)
    - Any descriptive label rows that happen to have numbers in D/E
    """
    children_by_region: dict[str, list[str]] = {r: [] for r in _TOP_REGIONS}
    current_region: str | None = None

    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, _COL_LABEL).value
        cagr = ws.cell(row, _COL_CAGR).value
        anchor = ws.cell(row, _COL_ANCHOR).value
        if not isinstance(name, str):
            continue
        name_stripped = name.strip()

        # Check if this is a region marker (exact or trailing-space variant)
        matched_region = None
        for region in _TOP_REGIONS:
            if name_stripped == region or name_stripped.rstrip() == region:
                matched_region = region
                break
        if matched_region:
            current_region = matched_region
            continue

        # Skip Global (rollup), dimension titles, descriptive labels, and
        # segmentation rows from non-geography dimensions.
        if name_stripped.lower() == "global":
            continue
        if _is_dimension_title(name_stripped):
            continue
        if name_stripped in segment_names and name_stripped not in region_segment_names:
            continue
        if _NON_GEO_PATTERNS.search(name_stripped):
            continue

        if (current_region
                and (isinstance(cagr, (int, float)) or isinstance(anchor, (int, float)))
                and name_stripped not in children_by_region[current_region]):
            children_by_region[current_region].append(name_stripped)

    region_nodes = tuple(
        GeographyNodeSchema(
            name=region,
            children=tuple(GeographyNodeSchema(name=c) for c in children),
        )
        for region, children in children_by_region.items()
        if children
    )
    return GeographyNodeSchema(name="Global", children=region_nodes)


def _extract_asp_products(ws_asp) -> list[str]:
    """Products from the ASP sheet.

    Primary strategy: find 'By Product' or 'Product' header and collect rows below it.
    Fallback: scan all rows with positive numeric values, skipping dimension headers,
    zero-value parent/rollup rows, and non-string labels. This handles markets
    where the ASP sheet uses a dimension header (e.g. 'By Laser Type') instead
    of 'By Product'.
    """
    label_col = find_label_column(ws_asp)
    products: list[str] = []
    in_block = False
    found_product_header = False

    for row in range(1, ws_asp.max_row + 1):
        label = ws_asp.cell(row, label_col).value
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label in ("By Product", "Product") or label.startswith("By Product"):
            if products:
                break          # stop at second block
            in_block = True
            found_product_header = True
            continue
        if in_block and _is_dimension_title(label):
            break
        if in_block and label:
            products.append(label)

    # Fallback: no 'By Product' header — scan all rows with positive values
    if not products:
        for row in range(1, ws_asp.max_row + 1):
            label = ws_asp.cell(row, label_col).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()
            if _is_dimension_title(label):
                continue
            if re.match(r"^\d+$", label):
                continue
            if label in ("Segment", "Geography", "Region"):
                continue
            # Check for at least one positive numeric value in columns to the right of label
            has_positive = any(
                isinstance(ws_asp.cell(row, c).value, (int, float))
                and ws_asp.cell(row, c).value > 0
                for c in range(label_col + 1, min(label_col + 8, _COL_GEO_END + 1))
            )
            if has_positive:
                products.append(label)

    return products


def _detect_priced_dimension(
    dimensions: list[DimensionSchema],
    asp_products: list[str],
) -> str:
    """Pick the dimension whose segments match ASP products."""
    asp_set = set(asp_products)
    for dim in dimensions:
        if set(dim.segments) == asp_set:
            return dim.title
    for dim in dimensions:
        if asp_set <= set(dim.segments):
            return dim.title
    if dimensions:
        return dimensions[0].title
    raise ValueError("No dimensions found in input workbook")


def extract_schema(input_path: Path | str) -> MarketSchema:
    """Extract market schema from an Input Sheet workbook."""
    path = Path(input_path)
    wb = load_workbook(path, data_only=True)

    data_sheet = find_data_sheet_name(wb)
    asp_sheet = find_asp_sheet_name(wb, exclude_sheet=data_sheet)
    if data_sheet is None:
        raise ValueError(f"'Data' sheet missing in {path}")
    if asp_sheet is None:
        raise ValueError(f"'ASP' sheet missing in {path}")

    ws_data = wb[data_sheet]
    ws_asp = wb[asp_sheet]

    market_name = str(ws_data.cell(*_MARKET_NAME_CELL).value or path.stem)
    dimensions = _extract_dimensions(ws_data)
    segment_names = frozenset(s for d in dimensions for s in d.segments)
    region_segment_names = frozenset(
        s for d in dimensions if _REGION_DIMENSION_PATTERN.search(d.title)
        for s in d.segments
    )
    geographies = _extract_geographies_from_headers(ws_data)
    geography_tree = _extract_geography_tree(
        ws_data, segment_names, region_segment_names)
    asp_products = _extract_asp_products(ws_asp)
    priced_dimension = _detect_priced_dimension(dimensions, asp_products)

    return MarketSchema(
        market_name=market_name,
        dimensions=tuple(dimensions),
        geographies=tuple(geographies),
        priced_dimension=priced_dimension,
        geography_tree=geography_tree,
    )


def infer_schema_path(input_path: Path | str, schemas_dir: Path | str = "schemas") -> Path:
    """Guess schema output path from the input filename."""
    stem = Path(input_path).stem.lower()
    # Strip common prefixes: "data file - ", "data file-", "input sheet - "
    slug = re.sub(r"^(data\s+file\s*[-–]\s*|input\s+sheet\s*[-–]\s*)", "", stem)
    slug = re.sub(r"\s+market\s*$", "", slug)
    slug = re.sub(r"[^a-z0-9]+", "_", slug).strip("_")
    return Path(schemas_dir) / f"{slug}.json"


def ensure_schema(
    input_path: Path | str,
    schema_path: Path | str | None = None,
) -> Path:
    """Return schema path, extracting to disk first if the file is missing."""
    path = Path(schema_path) if schema_path else infer_schema_path(input_path)
    if not path.exists():
        save_schema(extract_schema(input_path), path)
    return path


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Extract market schema from Input Sheet")
    p.add_argument("--input", required=True, help="Path to Input Sheet .xlsx")
    p.add_argument("--output", default=None, help="Output JSON path (auto-inferred if omitted)")
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else infer_schema_path(input_path)

    schema = extract_schema(input_path)
    save_schema(schema, output_path)
    print(f"Extracted schema for '{schema.market_name}'")
    print(f"  dimensions : {len(schema.dimensions)}")
    print(f"  geographies: {len(schema.geographies)}")
    print(f"  priced dim : {schema.priced_dimension}")
    print(f"  wrote      : {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
