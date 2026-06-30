"""WorkbookInspector — data-driven extraction of Input Sheet metadata.

Reads any Input Sheet workbook and extracts all available structural information
WITHOUT hardcoded assumptions about geography names, segment counts, or ASP rows.
The inspection result is saved to JSON for debugging and downstream consumption.

Usage (standalone):
    python -m me_engine.io.workbook_inspector --input "Data File - XYZ Market.xlsx"

Usage (library):
    from me_engine.io.workbook_inspector import WorkbookInspector
    meta = WorkbookInspector().inspect("Data File - XYZ Market.xlsx")
    meta.save("output/extracted_metadata/xyz_market.json")
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Column constants for the Input Sheet (1-based)
# These reflect the observed layout across all tested markets.
# ---------------------------------------------------------------------------
_COL_LABEL      = 2          # B — row label (segment names, geo names, etc.)
_COL_GEO_START  = 3          # C — first geography column in segmentation band
_COL_GEO_END    = 50         # AX — scan generously; we stop at the first blank row
_COL_CAGR       = 4          # D
_COL_ANCHOR     = 5          # E
_MARKET_NAME_CELL = (1, 3)   # C1

_DIMENSION_PREFIX = "By "

# Known top-level region names (used to build the geography tree).
# Extended vs extract_schema.py to cover more markets.
_KNOWN_REGIONS = frozenset({
    "North America", "Europe", "Asia Pacific", "Latin America",
    "Middle East", "Africa", "Middle East & Africa",
    "South America", "Central America", "Caribbean",
    "Eastern Europe", "Western Europe", "Southeast Asia", "Oceania",
    "Rest of the World",
})

# Labels that appear in the sizing block but are NOT geography names.
# These are header/label rows that happen to have numbers in CAGR/anchor columns.
_NON_GEO_LABEL_PATTERNS = re.compile(
    r"(market\s+size|us\$|usd|\$\s*mn|cagr|compound|forecast|estimate"
    r"|revenue|value|volume|note|source|base\s+year|currency|unit)",
    re.IGNORECASE,
)

_REGION_DIMENSION_PATTERN = re.compile(
    r"(region|geography|country|market)",
    re.IGNORECASE,
)


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.strip().lower())


def _is_dimension_title(label: str) -> bool:
    return label.startswith(_DIMENSION_PREFIX)


def _find_label_column(ws) -> int:
    """Return the best candidate column for row labels in an ASP-like sheet."""
    candidate_cols = (2, 3)
    scores: dict[int, tuple[int, int, int]] = {}
    max_row = min(ws.max_row, 120)

    for col in candidate_cols:
        strings = 0
        numerics = 0
        blanks = 0
        for row in range(1, max_row + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip():
                strings += 1
            elif isinstance(value, (int, float)):
                numerics += 1
            else:
                blanks += 1
        scores[col] = (strings, numerics, blanks)

    # Prefer the column with the most text labels; break ties toward the left.
    return max(candidate_cols, key=lambda c: (scores[c][0], -scores[c][1], -scores[c][2]))


def find_label_column(ws) -> int:
    """Public alias for the label column detector."""
    return _find_label_column(ws)


def _find_sheet_name(wb, expected_patterns: tuple[str, ...]) -> str | None:
    normalized = {
        _normalize_sheet_name(name): name
        for name in wb.sheetnames
    }
    for pattern in expected_patterns:
        key = _normalize_sheet_name(pattern)
        if key in normalized:
            return normalized[key]
    for norm, name in normalized.items():
        for pattern in expected_patterns:
            if _normalize_sheet_name(pattern) in norm:
                return name
    return None


def _sheet_has_data_structure(ws) -> bool:
    found_dim_header = False
    found_geo_header = False
    numeric_rows = 0
    for row in range(1, min(ws.max_row, 100) + 1):
        label = ws.cell(row, _COL_LABEL).value
        if isinstance(label, str) and label.strip().startswith(_DIMENSION_PREFIX):
            found_dim_header = True
        if row == 3:
            valid_headers = sum(
                1 for col in range(_COL_GEO_START, _COL_GEO_END + 1)
                if isinstance(ws.cell(row, col).value, str)
                and ws.cell(row, col).value.strip()
            )
            found_geo_header = valid_headers >= 2
        if isinstance(label, str) and label.strip() and label.strip() != "Global":
            cagr = ws.cell(row, _COL_CAGR).value
            anchor = ws.cell(row, _COL_ANCHOR).value
            if isinstance(cagr, (int, float)) or isinstance(anchor, (int, float)):
                numeric_rows += 1
    return found_dim_header and found_geo_header and numeric_rows >= 3


def _sheet_has_asp_structure(ws) -> bool:
    label_col = _find_label_column(ws)
    product_header = False
    positive_product_rows = 0
    candidate_rows = 0
    for row in range(1, min(ws.max_row, 200) + 1):
        label = ws.cell(row, label_col).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        if label == "Product" or label.startswith("By Product") or _is_dimension_title(label):
            product_header = True
            continue
        if label.startswith(_DIMENSION_PREFIX) and product_header and positive_product_rows >= 1:
            break
        if product_header:
            sample = [
                float(v) for col in range(max(_COL_GEO_START, label_col + 1), _COL_GEO_END + 1)
                if isinstance((v := ws.cell(row, col).value), (int, float))
            ]
            if sample:
                candidate_rows += 1
                if any(v > 0 for v in sample):
                    positive_product_rows += 1
        else:
            if label.startswith(_DIMENSION_PREFIX):
                continue
            if re.match(r"^\d+$", label):
                continue
            if label in ("Segment", "Geography") or _NON_GEO_LABEL_PATTERNS.search(label):
                continue
            sample = [
                float(v) for col in range(max(_COL_GEO_START, label_col + 1), _COL_GEO_END + 1)
                if isinstance((v := ws.cell(row, col).value), (int, float))
            ]
            if sample:
                candidate_rows += 1
                if any(v > 0 for v in sample):
                    positive_product_rows += 1

    if product_header:
        return positive_product_rows >= 1

    geo_headers = 0
    for col in range(max(_COL_GEO_START, label_col + 1), _COL_GEO_END + 1):
        header = ws.cell(3, col).value
        if isinstance(header, str) and header.strip():
            header_text = header.strip()
            if header_text not in ("Region", "Share", "CAGR", "Value"):
                geo_headers += 1
    return positive_product_rows >= 3 and geo_headers >= 4


def find_data_sheet_name(wb) -> str | None:
    explicit = _find_sheet_name(wb, ("Data", "Input", "InputSheet", "DataSheet"))
    if explicit:
        return explicit
    scored = [(sheet, _sheet_has_data_structure(wb[sheet])) for sheet in wb.sheetnames]
    for sheet, valid in scored:
        if valid:
            return sheet
    return None


def find_asp_sheet_name(wb, exclude_sheet: str | None = None) -> str | None:
    explicit = _find_sheet_name(
        wb,
        ("ASP", "ASPSheet", "Price", "AverageSellingPrice", "ASPPrice"),
    )
    if explicit:
        return explicit
    best_sheet = None
    for sheet in wb.sheetnames:
        if sheet == exclude_sheet:
            continue
        if _sheet_has_data_structure(wb[sheet]):
            continue
        if _sheet_has_asp_structure(wb[sheet]):
            best_sheet = sheet
            break
    if best_sheet is not None:
        return best_sheet
    if exclude_sheet is not None and _sheet_has_asp_structure(wb[exclude_sheet]):
        return exclude_sheet
    return None


def scan_asp_geo_columns(ws) -> dict[str, int]:
    """Return geography header columns for an ASP block.

    Some input workbooks reuse the same sheet for Data and ASP. In that case,
    row 3 contains repeated geography headers: the first occurrence belongs to
    the segmentation band, and later repeats belong to the ASP block.
    This helper maps each geography name to its last seen column index so the
    ASP reader/validator can choose the correct block when the sheet is shared.
    """
    label_col = _find_label_column(ws)
    geo_start = max(_COL_GEO_START, label_col + 1)

    occurrences: dict[str, list[int]] = {}
    for col in range(geo_start, _COL_GEO_END + 1):
        value = ws.cell(3, col).value
        if isinstance(value, str) and value.strip():
            occurrences.setdefault(value.strip(), []).append(col)

    if not occurrences:
        return {}

    product_rows: list[int] = []
    for row in range(1, min(ws.max_row, 200) + 1):
        label = ws.cell(row, label_col).value
        if not isinstance(label, str) or not label.strip():
            continue
        any_numeric = any(
            isinstance(ws.cell(row, c).value, (int, float))
            for c in range(geo_start, _COL_GEO_END + 1)
        )
        if any_numeric:
            product_rows.append(row)
        if len(product_rows) >= 20:
            break

    result: dict[str, int] = {}
    # For each geography, pick the occurrence column that has the most
    # numeric values across the sampled product rows. Fall back to the last
    # occurrence if no numeric evidence is found.
    for geo, cols in occurrences.items():
        best_col = None
        best_count = -1
        for col in cols:
            count = sum(
                1 for r in product_rows
                if isinstance(ws.cell(r, col).value, (int, float))
            )
            if count > best_count or (count == best_count and (best_col is None or col > best_col)):
                best_count = count
                best_col = col
        if best_col is None:
            best_col = cols[-1]
        result[geo] = best_col

    return result


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------

@dataclass
class DimensionInfo:
    """One segmentation dimension extracted from the Data sheet."""
    title: str
    segments: list[str]
    segments_found: int
    row_start: int    # 1-based row where dimension header was found
    row_end: int      # 1-based row of last segment


@dataclass
class GeographyInfo:
    """A geography node discovered in the sizing block (col B + CAGR/anchor)."""
    name: str
    cagr: float | None
    anchor_2025: float | None
    region: str | None
    has_cagr: bool
    has_anchor: bool
    has_column: bool  # whether it appears as a column header in the seg band


@dataclass
class AspProductInfo:
    """One product row found in the ASP sheet."""
    name: str
    row: int
    sample_values: list[float]   # first few numeric values from that row
    min_value: float | None
    max_value: float | None
    has_positive_values: bool


@dataclass
class ValidationIssue:
    """A single warning or failure found during inspection."""
    level: str          # "WARN" or "FAIL"
    category: str       # e.g. "missing-geography", "shares-sum", "asp-missing"
    detail: str

    def __str__(self) -> str:
        return f"[{self.level}] {self.category}: {self.detail}"


@dataclass
class WorkbookMetadata:
    """Complete structural metadata extracted from an Input Sheet workbook."""

    # --- Core identity ------------------------------------------------------
    market_name: str
    input_path: str
    sheet_names: list[str]

    # --- Data sheet ---------------------------------------------------------
    dimensions: list[DimensionInfo]
    geographies: list[GeographyInfo]     # from sizing block (col B)
    column_headers: list[str]            # geography names found in row 3

    # --- ASP sheet ----------------------------------------------------------
    asp_products: list[AspProductInfo]
    detected_priced_dimension: str | None

    # --- Numeric ranges (Data sheet, segmentation band) ---------------------
    share_value_range: dict[str, float]   # {"min": ..., "max": ...}

    # --- Computed summaries -------------------------------------------------
    missing_dimensions: list[str]         # dims where no segments were found
    missing_geographies: list[str]        # geo names in col B but no column
    extra_column_geos: list[str]          # cols in row 3 not in sizing block
    share_sum_issues: list[str]           # dims where shares ≠ 1 for some geos

    # --- Validation issues --------------------------------------------------
    issues: list[ValidationIssue]

    # --- Raw counts for quick scanning --------------------------------------
    total_dimensions: int
    total_segments: int
    total_geographies: int
    total_asp_products: int

    def passed(self) -> bool:
        return not any(i.level == "FAIL" for i in self.issues)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def save(self, path: Path | str) -> Path:
        """Write metadata to a JSON file, creating parent dirs as needed."""
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w", encoding="utf-8") as fh:
            json.dump(self.to_dict(), fh, indent=2, ensure_ascii=False)
            fh.write("\n")
        return p

    def summary_lines(self) -> list[str]:
        lines = [
            f"  Market       : {self.market_name}",
            f"  Sheet names  : {', '.join(self.sheet_names)}",
            f"  Dimensions   : {self.total_dimensions} "
            f"({', '.join(d.title for d in self.dimensions)})",
            f"  Segments     : {self.total_segments}",
            f"  Geographies  : {self.total_geographies} (col-B sizing block)",
            f"  Col headers  : {len(self.column_headers)} in seg band",
            f"  ASP products : {self.total_asp_products}",
            f"  Priced dim   : {self.detected_priced_dimension or '(unknown)'}",
        ]
        if self.missing_geographies:
            lines.append(f"  Missing cols : {', '.join(self.missing_geographies)}")
        if self.extra_column_geos:
            lines.append(f"  Extra cols   : {', '.join(self.extra_column_geos)}")
        if self.share_sum_issues:
            lines.append(f"  Share issues : {'; '.join(self.share_sum_issues)}")
        lines.append(f"  Issues       : "
                     f"{sum(1 for i in self.issues if i.level=='FAIL')} FAIL, "
                     f"{sum(1 for i in self.issues if i.level=='WARN')} WARN")
        return lines

    def __str__(self) -> str:
        return "\n".join(["=== WorkbookInspector Results ==="] + self.summary_lines()
                         + [str(i) for i in self.issues])



# ---------------------------------------------------------------------------
# Inspector
# ---------------------------------------------------------------------------

class WorkbookInspector:
    """Extracts all structural metadata from an Input Sheet workbook.

    The inspector is intentionally lenient: it collects everything it can
    find and records warnings/failures in the issues list rather than raising
    exceptions. Callers decide how to act on those issues.
    """

    def inspect(self, path: Path | str) -> WorkbookMetadata:
        """Main entry point. Returns a fully populated WorkbookMetadata."""
        input_path = Path(path)
        issues: list[ValidationIssue] = []

        # --- Open workbook --------------------------------------------------
        try:
            wb = load_workbook(input_path, data_only=True)
        except Exception as exc:
            return self._fatal(input_path, f"Cannot open workbook: {exc}")

        sheet_names = wb.sheetnames

        # --- Check required sheets ------------------------------------------
        data_sheet = find_data_sheet_name(wb)
        asp_sheet = find_asp_sheet_name(wb, exclude_sheet=data_sheet)

        if data_sheet is None:
            issues.append(ValidationIssue("FAIL", "sheet-exists", "'Data' sheet missing"))
        if asp_sheet is None:
            issues.append(ValidationIssue("FAIL", "sheet-exists", "'ASP' sheet missing"))

        if any(i.level == "FAIL" for i in issues):
            return self._empty(input_path, sheet_names, issues)

        if data_sheet != "Data":
            issues.append(ValidationIssue(
                "WARN", "sheet-name",
                f"Using '{data_sheet}' as the Data sheet"))
        if asp_sheet != "ASP":
            issues.append(ValidationIssue(
                "WARN", "sheet-name",
                f"Using '{asp_sheet}' as the ASP sheet"))

        ws_data = wb[data_sheet]
        ws_asp = wb[asp_sheet]

        # --- Extract market name --------------------------------------------
        market_name = self._extract_market_name(ws_data, input_path, issues)

        # --- Extract dimensions + segments from Data sheet ------------------
        dimensions = self._extract_dimensions(ws_data, issues)

        # --- Extract column headers (row 3, seg band) -----------------------
        column_headers = self._extract_column_headers(ws_data)

        # --- Extract geographies from sizing block (col B) ------------------
        geographies = self._extract_geographies(ws_data, column_headers, dimensions, issues)

        # --- Share sum checks -----------------------------------------------
        share_sum_issues = self._check_share_sums(ws_data, dimensions, geographies, issues)

        # --- Share value range ----------------------------------------------
        share_value_range = self._scan_share_value_range(ws_data, dimensions, geographies)

        # --- ASP sheet ------------------------------------------------------
        asp_products = self._extract_asp_products(ws_asp, issues)
        detected_priced_dim = self._detect_priced_dimension(dimensions, asp_products, issues)

        # --- Cross-reference: geo in col-B vs col headers -------------------
        col_header_set = set(column_headers)
        sizing_block_set = {g.name for g in geographies}
        missing_geos = [g.name for g in geographies if not g.has_column]
        extra_col_geos = [h for h in column_headers if h not in sizing_block_set]

        # Emit warnings for missing/extra cols (not failures)
        for name in missing_geos:
            issues.append(ValidationIssue(
                "WARN", "missing-geo-column",
                f"Geography '{name}' found in sizing block but has no column in seg band"))
        for name in extra_col_geos:
            issues.append(ValidationIssue(
                "WARN", "extra-column",
                f"Column header '{name}' in seg band not found in sizing block"))

        # FAIL if no usable geographies at all
        usable = [g for g in geographies if g.has_cagr and g.has_anchor]
        if not usable:
            issues.append(ValidationIssue(
                "FAIL", "no-usable-geographies",
                "No geographies with both CAGR and anchor value found"))

        if not column_headers:
            issues.append(ValidationIssue(
                "FAIL", "no-geo-columns",
                "No geography column headers found in row 3 of the segmentation band"))

        return WorkbookMetadata(
            market_name=market_name,
            input_path=str(input_path.resolve()),
            sheet_names=list(sheet_names),
            dimensions=dimensions,
            geographies=geographies,
            column_headers=column_headers,
            asp_products=asp_products,
            detected_priced_dimension=detected_priced_dim,
            share_value_range=share_value_range,
            missing_dimensions=[d.title for d in dimensions if d.segments_found == 0],
            missing_geographies=missing_geos,
            extra_column_geos=extra_col_geos,
            share_sum_issues=share_sum_issues,
            issues=issues,
            total_dimensions=len(dimensions),
            total_segments=sum(len(d.segments) for d in dimensions),
            total_geographies=len(geographies),
            total_asp_products=len(asp_products),
        )


    # --- private helpers ----------------------------------------------------

    @staticmethod
    def _fatal(path: Path, reason: str) -> WorkbookMetadata:
        """Return a minimal metadata object when the file cannot be opened."""
        return WorkbookMetadata(
            market_name="(unknown)",
            input_path=str(path),
            sheet_names=[],
            dimensions=[],
            geographies=[],
            column_headers=[],
            asp_products=[],
            detected_priced_dimension=None,
            share_value_range={},
            missing_dimensions=[],
            missing_geographies=[],
            extra_column_geos=[],
            share_sum_issues=[],
            issues=[ValidationIssue("FAIL", "file-open", reason)],
            total_dimensions=0,
            total_segments=0,
            total_geographies=0,
            total_asp_products=0,
        )

    @staticmethod
    def _empty(path: Path, sheet_names: list[str],
               issues: list[ValidationIssue]) -> WorkbookMetadata:
        """Return a minimal object when required sheets are missing."""
        return WorkbookMetadata(
            market_name="(unknown)",
            input_path=str(path),
            sheet_names=list(sheet_names),
            dimensions=[],
            geographies=[],
            column_headers=[],
            asp_products=[],
            detected_priced_dimension=None,
            share_value_range={},
            missing_dimensions=[],
            missing_geographies=[],
            extra_column_geos=[],
            share_sum_issues=[],
            issues=issues,
            total_dimensions=0,
            total_segments=0,
            total_geographies=0,
            total_asp_products=0,
        )

    @staticmethod
    def _extract_market_name(ws, path: Path,
                             issues: list[ValidationIssue]) -> str:
        val = ws.cell(*_MARKET_NAME_CELL).value
        if isinstance(val, str) and val.strip():
            return val.strip()
        issues.append(ValidationIssue(
            "WARN", "market-name", f"Cell C1 is blank or non-string (got {val!r}); using file stem '{path.stem}'"))
        return path.stem

    @staticmethod
    def _extract_dimensions(ws, issues: list[ValidationIssue]) -> list[DimensionInfo]:
        """Scan column B for 'By …' headers and collect segment rows beneath each."""
        dims: list[DimensionInfo] = []
        current_title: str | None = None
        current_segs: list[str] = []
        current_start: int = 0
        last_seg_row: int = 0

        for row in range(1, ws.max_row + 1):
            label = ws.cell(row, _COL_LABEL).value
            if not isinstance(label, str) or not label.strip():
                # Blank row — flush any in-progress dimension
                if current_title is not None and current_segs:
                    dims.append(DimensionInfo(
                        title=current_title,
                        segments=list(current_segs),
                        segments_found=len(current_segs),
                        row_start=current_start,
                        row_end=last_seg_row,
                    ))
                    current_title = None
                    current_segs = []
                continue

            label = label.strip()
            if label.startswith(_DIMENSION_PREFIX):
                # New dimension header found — flush previous if any
                if current_title is not None and current_segs:
                    dims.append(DimensionInfo(
                        title=current_title,
                        segments=list(current_segs),
                        segments_found=len(current_segs),
                        row_start=current_start,
                        row_end=last_seg_row,
                    ))
                current_title = label
                current_segs = []
                current_start = row
            elif current_title is not None:
                current_segs.append(label)
                last_seg_row = row

        # Flush last dimension
        if current_title is not None and current_segs:
            dims.append(DimensionInfo(
                title=current_title,
                segments=list(current_segs),
                segments_found=len(current_segs),
                row_start=current_start,
                row_end=last_seg_row,
            ))

        if not dims:
            issues.append(ValidationIssue(
                "FAIL", "no-dimensions",
                "No 'By …' dimension headers found in column B of Data sheet"))
        return dims


    @staticmethod
    def _extract_column_headers(ws) -> list[str]:
        """Geography names from row 3, scanning generously up to col _COL_GEO_END.

        Stops scanning once we hit two consecutive blank cells to avoid picking
        up stale headers far to the right of the actual data.
        """
        headers: list[str] = []
        seen: set[str] = set()
        blank_streak = 0

        for col in range(_COL_GEO_START, _COL_GEO_END + 1):
            val = ws.cell(3, col).value
            if not isinstance(val, str) or not val.strip():
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            name = val.strip()
            if name not in seen:
                headers.append(name)
                seen.add(name)

        return headers

    @staticmethod
    def _extract_geographies(ws, column_headers: list[str],
                              dimensions: list[DimensionInfo],
                              issues: list[ValidationIssue]) -> list[GeographyInfo]:
        """Extract geographies from the sizing block in column B.

        Looks for rows where:
        - col B is a non-blank string
        - col D (CAGR) or col E (anchor) contain a number
        - the label is NOT a dimension title and NOT a known non-geo label

        Also tracks which region each geo belongs to, using _KNOWN_REGIONS as
        region markers. Any region/country can appear — we infer the tree from order.
        """
        col_header_set = set(column_headers)
        geos: list[GeographyInfo] = []
        seen: set[str] = set()
        current_region: str | None = None

        segment_names = frozenset(s for d in dimensions for s in d.segments)
        region_rows = set()
        for dim in dimensions:
            if _REGION_DIMENSION_PATTERN.search(dim.title):
                region_rows.update(range(dim.row_start + 1, dim.row_end + 1))

        for row in range(1, ws.max_row + 1):
            name_val = ws.cell(row, _COL_LABEL).value
            cagr_val = ws.cell(row, _COL_CAGR).value
            anchor_val = ws.cell(row, _COL_ANCHOR).value

            if not isinstance(name_val, str) or not name_val.strip():
                continue

            name = name_val.strip()

            # Skip dimension headers and known segments outside a region/geography block
            if name.startswith(_DIMENSION_PREFIX):
                continue
            if name in segment_names and row not in region_rows:
                continue

            # Skip labels that look like descriptive text / column headers
            if _NON_GEO_LABEL_PATTERNS.search(name):
                continue

            # Detect region markers
            if name in _KNOWN_REGIONS:
                current_region = name
                continue

            # Require at least one numeric indicator (CAGR or anchor)
            has_cagr = isinstance(cagr_val, (int, float))
            has_anchor = isinstance(anchor_val, (int, float))
            if not (has_cagr or has_anchor):
                continue

            if name in seen:
                continue  # dedup (some sheets repeat)
            seen.add(name)

            cagr = float(cagr_val) if has_cagr else None
            anchor = float(anchor_val) if has_anchor else None

            if has_cagr and cagr <= 0:
                issues.append(ValidationIssue(
                    "WARN", "cagr-non-positive",
                    f"Geography '{name}' CAGR={cagr:.4f} is not positive"))
            if has_anchor and anchor <= 0:
                issues.append(ValidationIssue(
                    "WARN", "anchor-non-positive",
                    f"Geography '{name}' anchor={anchor} is not positive"))

            geos.append(GeographyInfo(
                name=name,
                cagr=cagr,
                anchor_2025=anchor,
                region=current_region,
                has_cagr=has_cagr,
                has_anchor=has_anchor,
                has_column=(name in col_header_set),
            ))

        return geos

    @staticmethod
    def _check_share_sums(ws, dimensions: list[DimensionInfo],
                          geographies: list[GeographyInfo],
                          issues: list[ValidationIssue]) -> list[str]:
        """Check that flat dimension shares sum to ~1.0 for each geography column.

        Only truly flat dimensions are checked (no parent relationships present).
        Hierarchical dimensions are skipped — checking their sum would require
        knowing the parent structure which we don't have without the schema.
        Returns a list of human-readable issue strings (also appended to issues).
        """
        _SHARE_TOL = 0.05  # allow ±5% rounding (lenient — just a warning)

        # Build a row map: label -> row number
        label_row: dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row, _COL_LABEL).value
            if isinstance(label, str) and label.strip():
                label_row[label.strip()] = row

        # Build col map from column_headers actually found
        geo_cols: dict[str, int] = {}
        for col in range(_COL_GEO_START, _COL_GEO_END + 1):
            val = ws.cell(3, col).value
            if isinstance(val, str) and val.strip() and val.strip() not in geo_cols:
                geo_cols[val.strip()] = col

        # Only flat dimensions: those whose segments are all present as label rows
        problems: list[str] = []
        sample_geos = [g.name for g in geographies
                       if g.has_column and g.has_cagr and g.has_anchor][:5]

        for dim in dimensions:
            seg_rows = [label_row.get(seg) for seg in dim.segments]
            if any(r is None for r in seg_rows):
                continue   # partial dimension — skip
            for geo in sample_geos:
                col = geo_cols.get(geo)
                if col is None:
                    continue
                total = sum(
                    float(ws.cell(seg_rows[i], col).value)
                    for i, _ in enumerate(dim.segments)
                    if isinstance(ws.cell(seg_rows[i], col).value, (int, float))
                )
                if total > 0 and abs(total - 1.0) > _SHARE_TOL:
                    msg = (f"'{dim.title}' shares for {geo} sum to {total:.4f} "
                           f"(expected ~1.0)")
                    problems.append(msg)
                    issues.append(ValidationIssue("WARN", "shares-sum", msg))
                    break  # one warning per dimension is enough

        return problems


    @staticmethod
    def _scan_share_value_range(ws, dimensions: list[DimensionInfo],
                                geographies: list[GeographyInfo]) -> dict[str, float]:
        """Find min/max of all share values in the segmentation band.

        Uses the first geography column and all segment rows to characterise
        what numeric values are present (should be 0–1 for shares).
        """
        values: list[float] = []

        # find first valid geo column
        first_col: int | None = None
        for col in range(_COL_GEO_START, _COL_GEO_END + 1):
            val = ws.cell(3, col).value
            if isinstance(val, str) and val.strip():
                first_col = col
                break

        if first_col is None:
            return {}

        for dim in dimensions:
            for row in range(dim.row_start + 1, dim.row_end + 1):
                v = ws.cell(row, first_col).value
                if isinstance(v, (int, float)):
                    values.append(float(v))

        if not values:
            return {}

        return {"min": min(values), "max": max(values),
                "mean": sum(values) / len(values), "count": len(values)}

    @staticmethod
    def _extract_asp_products(ws_asp, issues: list[ValidationIssue]) -> list[AspProductInfo]:
        """Extract all product rows from the ASP sheet.

        Strategy:
        1. Look for a 'By Product' or 'Product' header row (standard layout).
        2. If not found, fall back to collecting ALL rows with numeric values in
           nearby columns, treating any non-header, non-blank string in the
           detected label column as a product name. This handles markets where
           the ASP sheet uses a dimension header (e.g. 'By Laser Type') instead
           of 'By Product'.
        """
        products: list[AspProductInfo] = []
        in_product_block = False
        found_product_header = False
        label_col = find_label_column(ws_asp)

        for row in range(1, ws_asp.max_row + 1):
            label = ws_asp.cell(row, label_col).value
            if not isinstance(label, str) or not label.strip():
                continue

            label = label.strip()

            # 'By Product' header activates the product block
            if label in ("By Product", "Product"):
                in_product_block = True
                found_product_header = True
                continue

            # Any 'By …' header after the first block ends the block
            if label.startswith(_DIMENSION_PREFIX):
                if in_product_block and products:
                    break    # stop at the second "By …" block
                continue

            if not in_product_block:
                continue

            # Collect sample numeric values from columns to the right of the label column
            sample: list[float] = []
            for col in range(label_col + 1, min(label_col + 20, _COL_GEO_END + 1)):
                v = ws_asp.cell(row, col).value
                if isinstance(v, (int, float)):
                    sample.append(float(v))
                if len(sample) >= 8:
                    break

            if not sample:
                continue   # no numeric values → not a real product row

            has_positive = any(v > 0 for v in sample)
            if not has_positive:
                issues.append(ValidationIssue(
                    "WARN", "asp-non-positive",
                    f"ASP row for '{label}' has no positive values"))

            products.append(AspProductInfo(
                name=label,
                row=row,
                sample_values=sample[:5],
                min_value=min(sample) if sample else None,
                max_value=max(sample) if sample else None,
                has_positive_values=has_positive,
            ))

        # --- Fallback: scan ALL rows with numerics if no 'By Product' found --
        if not products and not found_product_header:
            for row in range(1, ws_asp.max_row + 1):
                label = ws_asp.cell(row, label_col).value
                if not isinstance(label, str) or not label.strip():
                    continue
                label = label.strip()
                if label.startswith(_DIMENSION_PREFIX):
                    continue
                # Skip pure-numeric or date-like labels (e.g. row 1 = '121')
                if re.match(r"^\d+$", label):
                    continue
                # Skip header/descriptor rows
                if label in ("Segment", "Geography") or _NON_GEO_LABEL_PATTERNS.search(label):
                    continue

                sample: list[float] = []
                for col in range(label_col + 1, min(label_col + 10, _COL_GEO_END + 1)):
                    v = ws_asp.cell(row, col).value
                    if isinstance(v, (int, float)):
                        sample.append(float(v))
                    if len(sample) >= 8:
                        break

                if not sample:
                    continue

                has_positive = any(v > 0 for v in sample)
                if not has_positive:
                    issues.append(ValidationIssue(
                        "WARN", "asp-non-positive",
                        f"ASP row for '{label}' has no positive values"))

                products.append(AspProductInfo(
                    name=label,
                    row=row,
                    sample_values=sample[:5],
                    min_value=min(sample) if sample else None,
                    max_value=max(sample) if sample else None,
                    has_positive_values=has_positive,
                ))

            if products:
                issues.append(ValidationIssue(
                    "WARN", "asp-no-product-header",
                    f"ASP sheet has no 'By Product' header; "
                    f"collected {len(products)} rows by numeric scan (fallback)"))

        if not products:
            issues.append(ValidationIssue(
                "FAIL", "no-asp-products",
                "No product rows found in ASP sheet under 'By Product' header"))

        return products

    @staticmethod
    def _detect_priced_dimension(dimensions: list[DimensionInfo],
                                  asp_products: list[AspProductInfo],
                                  issues: list[ValidationIssue]) -> str | None:
        """Pick the dimension whose segment set best matches the ASP products."""
        if not asp_products:
            return None

        asp_names = {p.name for p in asp_products}

        # Exact match first
        for dim in dimensions:
            if set(dim.segments) == asp_names:
                return dim.title

        # Subset match (ASP products ⊆ dimension segments)
        for dim in dimensions:
            if asp_names <= set(dim.segments):
                return dim.title

        # Overlap match (best overlap)
        best_dim = max(
            dimensions,
            key=lambda d: len(asp_names & set(d.segments)),
            default=None,
        )
        if best_dim and len(asp_names & set(best_dim.segments)) > 0:
            overlap = asp_names & set(best_dim.segments)
            issues.append(ValidationIssue(
                "WARN", "priced-dim-partial-match",
                f"Priced dimension '{best_dim.title}' matches only "
                f"{len(overlap)}/{len(asp_names)} ASP products"))
            return best_dim.title

        if dimensions:
            issues.append(ValidationIssue(
                "WARN", "priced-dim-unknown",
                "Could not match ASP products to any dimension — defaulting to first"))
            return dimensions[0].title

        return None



# ---------------------------------------------------------------------------
# Metadata path helper
# ---------------------------------------------------------------------------

def metadata_output_path(input_path: Path | str,
                          out_dir: Path | str = "output/extracted_metadata") -> Path:
    """Derive a canonical JSON path for a given input workbook."""
    stem = Path(input_path).stem.lower()
    slug = re.sub(r"^data\s+file\s*[-–—]\s*", "", stem)   # strip "Data File - " prefix
    slug = re.sub(r"\s+market\s*$", "", slug)               # strip trailing "Market"
    slug = re.sub(r"[^a-z0-9]+", "_", slug).strip("_")
    return Path(out_dir) / f"{slug}_metadata.json"


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _parse_args():
    import argparse
    p = argparse.ArgumentParser(description="Inspect an Input Sheet workbook")
    p.add_argument("--input", required=True, help="Path to Input Sheet .xlsx")
    p.add_argument("--output", default=None, help="Output JSON path (auto-inferred if omitted)")
    return p.parse_args()


def main() -> int:
    import sys
    args = _parse_args()
    input_path = Path(args.input)
    out_path = Path(args.output) if args.output else metadata_output_path(input_path)

    meta = WorkbookInspector().inspect(input_path)
    saved = meta.save(out_path)

    print(str(meta))
    print(f"\n  Saved → {saved}")
    return 0 if meta.passed() else 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
