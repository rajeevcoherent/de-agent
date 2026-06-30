"""Read the Input Sheet into per-geography curve context.

The Input Sheet is the agent's starting point: it supplies the forecast CAGR, the
2025 anchor value, and the region each geography belongs to. (It does NOT contain
the growth curve — that is exactly what the Curve Agent produces.)
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..curve.agent import MarketContext
from ..domain.runtime_taxonomy import RuntimeTaxonomy, default_taxonomy
from .workbook_inspector import find_data_sheet_name

_NAME_COL, _CAGR_COL, _VALUE_COL = 2, 4, 5      # cols B, D, E
_MARKET_NAME_CELL = (1, 3)                      # C1


@dataclass(frozen=True, slots=True)
class GeoInput:
    name: str
    forecast_cagr: float
    anchor_2025: float
    region: str | None


class InputSheetReader:
    """Extracts geography CAGR/value/region rows from the Input Sheet."""

    def __init__(self, path: Path | str, sheet: str | None = None,
                 taxonomy: RuntimeTaxonomy | None = None) -> None:
        self._taxonomy = taxonomy or default_taxonomy()
        wb = load_workbook(Path(path), data_only=True)
        sheet_name = sheet or find_data_sheet_name(wb)
        if sheet_name is None:
            raise ValueError(
                "Input workbook is missing a Data sheet; expected Data/Input/InputSheet")
        self._ws = wb[sheet_name]
        raw_market_name = self._ws.cell(*_MARKET_NAME_CELL).value
        self._market_name = (
            str(raw_market_name).strip()
            if isinstance(raw_market_name, str) and raw_market_name.strip()
            else Path(path).stem
        )

    def market_name(self) -> str:
        return self._market_name

    def read(self) -> dict[str, GeoInput]:
        """Geography rows from the sizing block, keyed by name.

        Only rows whose label is a *known geography* are kept — this excludes the
        segmentation rows (product/packaging/etc.) at the top of the sheet that
        also carry numbers. The sizing block lists each geography with its CAGR
        and 2025 value; the last occurrence wins (dedupes the two sub-blocks).
        """
        valid_names = set(self._taxonomy.geographies.by_name)
        parent_of = self._taxonomy.geographies.parent_of
        found: dict[str, GeoInput] = {}
        for row in range(1, self._ws.max_row + 1):
            name = self._ws.cell(row, _NAME_COL).value
            cagr = self._ws.cell(row, _CAGR_COL).value
            value = self._ws.cell(row, _VALUE_COL).value
            if (isinstance(name, str) and name in valid_names
                    and isinstance(cagr, (int, float))
                    and isinstance(value, (int, float))):
                found[name] = GeoInput(
                    name=name,
                    forecast_cagr=float(cagr),
                    anchor_2025=float(value),
                    region=parent_of.get(name),
                )
        return found

    def contexts(self) -> list[MarketContext]:
        market = self.market_name()
        return [
            MarketContext(market_name=market, geography=g.name,
                          forecast_cagr=g.forecast_cagr, region=g.region)
            for g in self.read().values()
        ]
