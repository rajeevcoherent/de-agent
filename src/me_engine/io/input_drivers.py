"""Build a DriverSet from the Input Sheet + agent curves (the pilot data path).

This is the live-system analogue of MEWorkbookReader: instead of reading finished
dynamics out of an ME file, it composes them from
  - the Input Sheet's single-year snapshots (segmentation %, ASP), and
  - the Curve Agent's value path per geography.

Where the Input Sheet has only a single year, we hold that value flat across the
horizon. This is an explicit, measurable assumption: the gap it creates versus the
human ME file quantifies exactly how much the (not-yet-built) segmentation/ASP
drift agents still need to supply.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..curve.asp_agent import AspAgent
from ..curve.runner import CurveRunner, truth_paths_from_me
from .workbook_inspector import (
    find_data_sheet_name,
    find_asp_sheet_name,
    find_label_column,
    scan_asp_geo_columns,
)
from ..curve.seg_agent import SegmentationAgent
from ..curve.segmentation import ShareDriftBuilder
from ..domain.drivers import (
    AspDrivers, DriverSet, GeographyDrivers, SegmentationDrivers,
)
from ..domain.series import Series
from ..domain.runtime_taxonomy import RuntimeTaxonomy, default_taxonomy
from ..domain.taxonomy import BASE_YEAR, YEARS
from .input_reader import InputSheetReader


def _flat(value: float) -> Series:
    """Hold a single-year snapshot flat across the whole horizon."""
    return Series({year: value for year in YEARS})


class InputDriverBuilder:
    """Composes Input-Sheet snapshots + agent curves into a DriverSet."""

    def __init__(self, input_path: Path | str,
                 truth_me: Path | str | None = None,
                 use_segmentation_agent: bool = True,
                 taxonomy: RuntimeTaxonomy | None = None) -> None:
        self._taxonomy = taxonomy or default_taxonomy()
        self._input_path = Path(input_path)
        wb = load_workbook(self._input_path, data_only=True)
        data_sheet = find_data_sheet_name(wb)
        asp_sheet = find_asp_sheet_name(wb, exclude_sheet=data_sheet)
        if data_sheet is None or asp_sheet is None:
            raise ValueError(
                "Input workbook must contain Data and ASP sheets (or equivalent names)")
        self._data = wb[data_sheet]
        self._asp = wb[asp_sheet]
        self._truth = truth_paths_from_me(truth_me) if truth_me else None
        self._col_of = self._scan_country_columns()
        self._asp_col_of = self._scan_asp_country_columns()
        self._drift = ShareDriftBuilder()
        market = InputSheetReader(self._input_path, taxonomy=self._taxonomy).market_name()
        self._premiums = (self._decide_premiums(market)
                          if use_segmentation_agent else {})
        self._asp_agent = AspAgent()
        # Pass only the leaf products to the ASP agent (parent rollup rows are
        # not real products and have no meaningful ASP values).
        priced = self._taxonomy.priced_dimension
        parents_in_priced = {
            priced.parent_of(s) for s in priced.segments if priced.parent_of(s) is not None
        }
        from ..domain.taxonomy import Dimension as _Dim
        leaf_priced = _Dim(
            title=priced.title,
            segments=tuple(s for s in priced.segments if s not in parents_in_priced),
            parent={s: priced.parent_of(s) for s in priced.segments
                    if s not in parents_in_priced},
        )
        self._asp_decision = (
            self._asp_agent.decide(market, leaf_priced)
            if use_segmentation_agent else None
        )

    def _decide_premiums(self, market: str) -> dict:
        """Decide per-segment growth premiums once per market (structural)."""
        agent = SegmentationAgent()
        return {dim.title: agent.decide(market, dim).premiums
                for dim in self._taxonomy.segmentation_dimensions}

    def build(self) -> DriverSet:
        reader = InputSheetReader(self._input_path, taxonomy=self._taxonomy)
        inputs = reader.read()
        self._curve_outcomes = {
            o.geo.name: o for o in CurveRunner().run(
                self._input_path, self._truth, taxonomy=self._taxonomy)
        }

        geographies = {}
        for name, geo in inputs.items():
            if name not in self._col_of:
                continue                          # only geos with a segmentation column
            value = (self._curve_outcomes[name].path
                     if name in self._curve_outcomes else _flat(geo.anchor_2025))
            geographies[name] = GeographyDrivers(
                name=name,
                value=value,
                segmentation=self._segmentation(name, geo.forecast_cagr),
                asp=self._asp_drivers(name),
            )
        return DriverSet(market_name=reader.market_name(), geographies=geographies)

    def curve_rationale(self) -> dict[str, str]:
        """Return per-geography curve agent reasoning after build() is called."""
        if not hasattr(self, "_curve_outcomes"):
            return {}
        rationale = {}
        for name, outcome in self._curve_outcomes.items():
            d = outcome.decision
            tag = "FALLBACK" if d.used_fallback else "AGENT"
            score_str = ""
            if outcome.score:
                sc = outcome.score
                score_str = (f" | shape_err={sc.shape_mae:.4f} "
                             f"cosine={sc.cosine:.4f} path_max={sc.path_max_rel_error:.2%}")
            rationale[name] = (
                f"[{tag}] archetype={d.archetype} peak={d.peak_year} "
                f"conf={d.confidence:.2f}{score_str} | {d.reasoning}"
            )
        return rationale

    # --- segmentation & asp from single-year snapshots ----------------------
    def _segmentation(self, geo: str, market_cagr: float) -> SegmentationDrivers:
        """Base-year shares from the Input Sheet, drifted by agent premiums.

        For hierarchical dimensions, premiums are only applied to leaf (child)
        segments — parent segments' shares are derived from their children and
        should not have independent premiums, as that would double-count drift.

        When no premiums are available the drift is zero and shares stay flat.
        """
        col = self._col_of[geo]
        shares: dict[str, dict[str, Series]] = {}
        for dim in self._taxonomy.segmentation_dimensions:
            base = self._read_base_shares(dim, col)
            raw_premiums = self._premiums.get(dim.title, {})

            # For hierarchical dims: strip premiums from parent segments.
            # Parents' values emerge from children; an independent premium on a
            # parent would compound with the children's premiums and over-inflate.
            has_any_children = any(dim.parent_of(s) is not None for s in dim.segments)
            if has_any_children and raw_premiums:
                parent_names = {dim.parent_of(s) for s in dim.segments
                                if dim.parent_of(s) is not None}
                premiums = {seg: p for seg, p in raw_premiums.items()
                            if seg not in parent_names}
            else:
                premiums = raw_premiums

            if premiums:
                shares[dim.title] = self._drift.build(dim, base, market_cagr, premiums)
            else:
                shares[dim.title] = {seg: _flat(v) for seg, v in base.items()}
        return SegmentationDrivers(shares=shares)

    def _asp_drivers(self, geo: str) -> AspDrivers:
        """Base-year ASP per product, grown by the agent's inflation rate.

        Only leaf segments of the priced dimension are processed — parent/rollup
        rows (those that have children) are skipped because their ASP cell is
        either zero or a sum, not a real price.

        With no ASP decision (segmentation agent disabled) prices stay flat,
        recovering the earlier baseline.
        """
        col = self._col_of[geo]
        dim = self._taxonomy.priced_dimension

        # Identify parent segments (those that have at least one child)
        parents_in_dim = {
            dim.parent_of(s) for s in dim.segments if dim.parent_of(s) is not None
        }
        leaf_products = [p for p in dim.segments if p not in parents_in_dim]

        asp: dict[str, Series] = {}
        for product in leaf_products:
            asp_col = self._asp_col_of.get(geo, col)
            try:
                base = self._asp_cell(self._asp_row(product), asp_col)
            except (KeyError, TypeError, ValueError):
                base = 0.0
            if self._asp_decision is not None:
                rate = self._asp_decision.rates.get(product,
                       self._asp_decision.rates.get(list(self._asp_decision.rates)[0], 0.0085))
                asp[product] = self._asp_agent.price_path(base, rate)
            else:
                asp[product] = _flat(base)
        return AspDrivers(asp=asp)

    def _read_base_shares(self, dim, col: int) -> dict[str, float]:
        """Read base-year shares for one dimension, inferring None cells.

        Some input sheets store only N-1 of N siblings — the last/first is
        left blank and implied as the complement (1 - sum of siblings with the
        same parent). This is common for the Distribution Channel hierarchy.
        """
        raw: dict[str, float | None] = {}
        for seg in dim.segments:
            v = self._data.cell(self._segment_row(dim, seg), col).value
            raw[seg] = float(v) if isinstance(v, (int, float)) else None

        # Group siblings by parent and infer any single None as complement.
        from collections import defaultdict
        by_parent: dict[str | None, list[str]] = defaultdict(list)
        for seg in dim.segments:
            by_parent[dim.parent_of(seg)].append(seg)

        result: dict[str, float] = {}
        for parent, siblings in by_parent.items():
            nones = [s for s in siblings if raw[s] is None]
            known = [s for s in siblings if raw[s] is not None]
            if len(nones) == 0:
                for s in siblings:
                    result[s] = raw[s]
            elif len(nones) == 1:
                # Infer missing sibling as complement of the rest
                known_sum = sum(raw[s] for s in known)
                result[nones[0]] = max(0.0, 1.0 - known_sum)
                for s in known:
                    result[s] = raw[s]
            else:
                # Multiple Nones: fall back to equal split of remainder
                known_sum = sum(raw[s] for s in known) if known else 0.0
                remainder = max(0.0, 1.0 - known_sum)
                share_each = remainder / len(nones)
                for s in known:
                    result[s] = raw[s]
                for s in nones:
                    result[s] = share_each

        return result

    # --- sheet scanning helpers --------------------------------------------
    # The segmentation %% band occupies the first country header band only; the
    # sheet repeats country headers further right for CAGR and trend-ID bands.
    # Use 50 cols (matching WorkbookInspector) so wide markets aren't truncated.
    _SEGMENTATION_BAND_END_COL = 50      # column AX (generous upper bound)

    def _scan_country_columns(self) -> dict[str, int]:
        """Map geography name -> its column in the first segmentation band.

        Scans cols C–AX, stopping at two consecutive blank headers to avoid
        picking up repeated geo-header bands (CAGR, trend-ID, etc.).
        Only geographies that appear in the taxonomy are included.
        """
        valid = set(self._taxonomy.geographies.by_name)
        result: dict[str, int] = {}
        blank_streak = 0
        for c in range(3, self._SEGMENTATION_BAND_END_COL + 1):
            val = self._data.cell(3, c).value
            if not isinstance(val, str) or not val.strip():
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            name = val.strip()
            if name in valid and name not in result:
                result[name] = c
        return result

    def _scan_asp_country_columns(self) -> dict[str, int]:
        """Map geography name -> its ASP column in the ASP sheet.

        If the ASP sheet is the same as the Data sheet, this finds the last
        occurrence of each geography name in row 3 so the ASP block is used
        instead of the first segmentation band.
        """
        # Use the inspector's smarter scan which picks the occurrence column
        # that contains numeric ASP values for product rows when the Data and
        # ASP blocks share the same sheet.
        cols = scan_asp_geo_columns(self._asp)
        # Filter to only the geographies present in the taxonomy
        valid = set(self._taxonomy.geographies.by_name)
        return {name: col for name, col in cols.items() if name in valid}

    def _segment_row(self, dim, segment: str) -> int:
        return self._find_row(self._data, segment)

    def _asp_row(self, product: str) -> int:
        return self._find_row(self._asp, product)

    @staticmethod
    def _find_row(ws, label: str) -> int:
        label_col = find_label_column(ws)
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(r, label_col).value
            if isinstance(cell_val, str) and cell_val.strip() == label:
                return r
        for r in range(1, ws.max_row + 1):
            for c in (2, 3, 4):
                cell_val = ws.cell(r, c).value
                if isinstance(cell_val, str) and cell_val.strip() == label:
                    return r
        raise KeyError(f"label not found: {label}")

    def _cell(self, row: int, col: int) -> float:
        return float(self._data.cell(row, col).value)

    def _asp_cell(self, row: int, col: int) -> float:
        return float(self._asp.cell(row, col).value)
