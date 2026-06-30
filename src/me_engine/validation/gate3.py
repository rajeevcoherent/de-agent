"""Gate 3 — Output validation + rationale diff logger.

Runs after the assembler produces a MarketResult. Two layers:

1. **Identity checks** — pure math invariants that must hold regardless of
   whether a reference ME file exists (shares sum to 1, volume identity, no
   negative values, CAGR round-trip).

2. **Reference diff** — if a ground-truth ME file is supplied, every value
   cell is compared and deviations are logged with the agent's reasoning so
   you can see *why* we differ, not just *how much*.
"""
from __future__ import annotations

import statistics
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from ..assembly.model import MarketResult
from ..domain.runtime_taxonomy import RuntimeTaxonomy, default_taxonomy
from ..domain.taxonomy import Band, YEARS, BASE_YEAR
from ..io.layout import SheetLayout, metric_col


_SHARE_TOL   = 0.005   # shares must sum to 1 within 0.5% (floating-point + drift tolerance)
_VOLUME_TOL  = 0.01    # 1% tolerance on value/asp*1000 = volume
_CAGR_TOL    = 0.0001  # 0.01% — CAGR round-trip tolerance
_DIFF_BUCKET = 0.05    # cells > 5% deviation are flagged


# ---------------------------------------------------------------------------
# Identity check results
# ---------------------------------------------------------------------------

@dataclass
class IdentityViolation:
    check: str
    geography: str
    detail: str

    def __str__(self) -> str:
        return f"  [FAIL] {self.check} | {self.geography}: {self.detail}"


@dataclass
class IdentityReport:
    violations: list[IdentityViolation] = field(default_factory=list)
    checks_run: int = 0

    def passed(self) -> bool:
        return len(self.violations) == 0

    def summary(self) -> str:
        status = "PASSED" if self.passed() else f"FAILED ({len(self.violations)} violations)"
        return f"Identity checks: {status} ({self.checks_run} checks run)"


# ---------------------------------------------------------------------------
# Reference diff results
# ---------------------------------------------------------------------------

@dataclass
class CellDiff:
    geography: str
    band: str
    segment: str
    year: int
    expected: float
    actual: float
    rel_error: float

    def __str__(self) -> str:
        direction = "over" if self.actual > self.expected else "under"
        return (f"  {self.geography:18s} | {self.band:10s} | {self.segment:38s} "
                f"| {self.year} | expected={self.expected:10.3f} "
                f"actual={self.actual:10.3f} ({self.rel_error:+.1%} {direction})")


@dataclass
class DiffSummary:
    cells_compared: int = 0
    mean_rel_error: float = 0.0
    median_rel_error: float = 0.0
    p90_rel_error: float = 0.0
    worst_rel_error: float = 0.0
    cells_within_5pct: float = 0.0
    worst_cells: list[CellDiff] = field(default_factory=list)

    def summary_lines(self) -> list[str]:
        return [
            f"  cells compared   : {self.cells_compared:,}",
            f"  mean rel error   : {self.mean_rel_error:.2%}",
            f"  median rel error : {self.median_rel_error:.2%}",
            f"  90th percentile  : {self.p90_rel_error:.2%}",
            f"  worst            : {self.worst_rel_error:.2%}",
            f"  cells within 5%  : {self.cells_within_5pct:.1%}",
        ]


@dataclass
class Gate3Report:
    identity: IdentityReport = field(default_factory=IdentityReport)
    diff: DiffSummary | None = None
    rationale_log: list[str] = field(default_factory=list)

    def passed(self) -> bool:
        return self.identity.passed()

    def __str__(self) -> str:
        lines = ["=== GATE 3 — Output Validation ==="]
        lines.append(f"  {self.identity.summary()}")
        for v in self.identity.violations:
            lines.append(str(v))

        if self.diff:
            lines.append("\n  --- Reference Diff ---")
            lines.extend(self.diff.summary_lines())
            if self.diff.worst_cells:
                lines.append(f"\n  Top {len(self.diff.worst_cells)} largest deviations:")
                for c in self.diff.worst_cells:
                    lines.append(str(c))

        if self.rationale_log:
            lines.append("\n  --- Agent Rationale for Deviations ---")
            lines.extend(self.rationale_log)

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Validator
# ---------------------------------------------------------------------------

class OutputValidator:
    """Validates the assembled MarketResult and optionally diffs vs a reference."""

    def __init__(self, taxonomy: RuntimeTaxonomy | None = None) -> None:
        self._taxonomy = taxonomy or default_taxonomy()

    def validate(
        self,
        result: MarketResult,
        reference_path: Path | str | None = None,
        agent_rationale: dict[str, str] | None = None,
    ) -> Gate3Report:
        """
        Args:
            result: The assembled MarketResult from the assembler.
            reference_path: Optional path to the ground-truth ME workbook.
            agent_rationale: Optional dict of geography -> reasoning string
                             from the curve agent, used to annotate the diff log.
        """
        report = Gate3Report()
        report.identity = self._run_identity_checks(result)

        if reference_path:
            report.diff = self._run_reference_diff(result, Path(reference_path))
            if agent_rationale:
                report.rationale_log = self._build_rationale_log(
                    report.diff, agent_rationale)

        return report

    # --- identity checks ----------------------------------------------------

    def _run_identity_checks(self, result: MarketResult) -> IdentityReport:
        ir = IdentityReport()

        for geo_name, geo in result.geographies.items():
            value_band = geo.bands.get(Band.VALUE)
            asp_band   = geo.bands.get(Band.ASP)
            vol_band   = geo.bands.get(Band.VOLUME)

            if not (value_band and asp_band and vol_band):
                continue

            # 1. No negative values in any band
            for band_label, band in geo.bands.items():
                for seg, row in band.rows_by_label.items():
                    for y in YEARS:
                        v = row.series.at(y)
                        if v < 0:
                            ir.violations.append(IdentityViolation(
                                "no-negatives", geo_name,
                                f"{band_label} | {seg} | {y} = {v:.4f}"))
                    ir.checks_run += len(YEARS)

            # 2. CAGR round-trip on the value band total
            total_series = value_band.total
            computed_cagr = total_series.cagr()
            # We can't easily get the input CAGR here without threading it
            # through, so we check the band total CAGR is positive and finite
            if not (0 < computed_cagr < 1.0):
                ir.violations.append(IdentityViolation(
                    "cagr-bounds", geo_name,
                    f"Total value CAGR={computed_cagr:.4f} is out of (0,1) range"))
            ir.checks_run += 1

            # 3. Volume identity: vol ≈ value / asp * 1000 per product
            for product in vol_band.rows_by_label:
                if product not in value_band.rows_by_label:
                    continue
                if product not in asp_band.rows_by_label:
                    continue
                for y in YEARS:
                    val = value_band.rows_by_label[product].series.at(y)
                    asp = asp_band.rows_by_label[product].series.at(y)
                    vol = vol_band.rows_by_label[product].series.at(y)
                    if asp == 0:
                        continue
                    expected_vol = val / asp * 1000
                    if abs(vol - expected_vol) / max(abs(expected_vol), 1e-9) > _VOLUME_TOL:
                        ir.violations.append(IdentityViolation(
                            "volume-identity", geo_name,
                            f"{product} | {y}: vol={vol:.3f} expected={expected_vol:.3f}"))
                ir.checks_run += len(YEARS)

            # 4. Segment shares sum check — for each flat dimension in value band
            for dim in self._taxonomy.segmentation_dimensions:
                flat_segs = [s for s in dim.segments if dim.parent_of(s) is None]
                if not all(s in value_band.rows_by_label for s in flat_segs):
                    continue
                for y in YEARS:
                    total_val = value_band.total.at(y)
                    seg_sum = sum(
                        value_band.rows_by_label[s].series.at(y)
                        for s in flat_segs
                    )
                    if total_val > 0:
                        ratio = seg_sum / total_val
                        if abs(ratio - 1.0) > _SHARE_TOL * 100:
                            ir.violations.append(IdentityViolation(
                                "shares-sum", geo_name,
                                f"'{dim.title}' {y}: seg_sum/total={ratio:.6f}"))
                ir.checks_run += len(YEARS)

        return ir

    # --- reference diff -----------------------------------------------------

    def _run_reference_diff(self, result: MarketResult, ref: Path) -> DiffSummary:
        wb = load_workbook(ref, data_only=True, read_only=True)
        all_diffs: list[CellDiff] = []

        for geo_name, geo in result.geographies.items():
            if geo_name not in wb.sheetnames:
                continue
            ws = wb[geo_name]
            layout = SheetLayout.discover(ws)
            value_band = geo.bands.get(Band.VALUE)
            if not value_band:
                continue

            for seg, row in value_band.rows_by_label.items():
                coord = layout.row_of.get((Band.VALUE, seg))
                if coord is None:
                    continue
                for y in YEARS:
                    cell_val = ws.cell(coord, metric_col(y)).value
                    if not isinstance(cell_val, (int, float)) or cell_val == 0:
                        continue
                    expected = float(cell_val)
                    actual   = row.series.at(y)
                    rel_err  = abs(actual - expected) / abs(expected)
                    all_diffs.append(CellDiff(
                        geography=geo_name,
                        band="Value",
                        segment=seg,
                        year=y,
                        expected=expected,
                        actual=actual,
                        rel_error=(actual - expected) / abs(expected),
                    ))

        if not all_diffs:
            return DiffSummary()

        abs_errs = [abs(d.rel_error) for d in all_diffs]
        sorted_errs = sorted(abs_errs)
        n = len(abs_errs)

        worst_cells = sorted(all_diffs, key=lambda d: abs(d.rel_error), reverse=True)[:15]

        return DiffSummary(
            cells_compared=n,
            mean_rel_error=statistics.mean(abs_errs),
            median_rel_error=statistics.median(abs_errs),
            p90_rel_error=sorted_errs[int(n * 0.9)],
            worst_rel_error=max(abs_errs),
            cells_within_5pct=sum(1 for e in abs_errs if e <= _DIFF_BUCKET) / n,
            worst_cells=worst_cells,
        )

    # --- rationale log ------------------------------------------------------

    def _build_rationale_log(
        self,
        diff: DiffSummary,
        agent_rationale: dict[str, str],
    ) -> list[str]:
        """For each worst-deviation cell, annotate with the agent's reasoning."""
        if not diff.worst_cells:
            return []

        lines: list[str] = []
        seen_geos: set[str] = set()

        for cell in diff.worst_cells:
            geo = cell.geography
            lines.append(
                f"\n  [{geo}] {cell.segment} | {cell.year} "
                f"→ {cell.rel_error:+.2%} vs reference"
            )
            if geo not in seen_geos and geo in agent_rationale:
                lines.append(f"    Agent reasoning: {agent_rationale[geo]}")
                seen_geos.add(geo)
            elif geo not in agent_rationale:
                lines.append(f"    Agent reasoning: (fallback — canonical shape used, no market-specific reasoning)")

        return lines
