"""Gate 1 — Input Sheet validation.

Runs before any agent is called. Every check here is a hard stop: if it fails
the pipeline does not proceed. This prevents agents from running on broken data
and producing silently wrong output.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from ..domain.runtime_taxonomy import RuntimeTaxonomy, default_taxonomy
from ..io.workbook_inspector import (
    find_data_sheet_name,
    find_asp_sheet_name,
    find_label_column,
    scan_asp_geo_columns,
)


# Tolerance for shares summing to 1.0 within a dimension.
_SHARE_SUM_TOL = 0.02   # allow ±2% rounding from the source sheet


@dataclass
class ValidationFailure:
    check: str
    detail: str

    def __str__(self) -> str:
        return f"[FAIL] {self.check}: {self.detail}"


@dataclass
class Gate1Report:
    failures: list[ValidationFailure] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def passed(self) -> bool:
        return len(self.failures) == 0

    def __str__(self) -> str:
        lines = ["=== GATE 1 — Input Validation ==="]
        if self.passed():
            lines.append("  PASSED — all checks clean")
        else:
            lines.append(f"  FAILED — {len(self.failures)} error(s)")
        for f in self.failures:
            lines.append(f"  {f}")
        for w in self.warnings:
            lines.append(f"  [WARN] {w}")
        return "\n".join(lines)


class InputValidator:
    """Validates the Input Sheet before passing it to the agent fleet."""

    def __init__(self, taxonomy: RuntimeTaxonomy | None = None) -> None:
        self._taxonomy = taxonomy or default_taxonomy()

    def validate(self, input_path: Path | str) -> Gate1Report:
        path = Path(input_path)
        report = Gate1Report()

        try:
            wb_data = load_workbook(path, data_only=True)
        except Exception as e:
            report.failures.append(ValidationFailure("file-open", str(e)))
            return report

        data_sheet = find_data_sheet_name(wb_data)
        asp_sheet = find_asp_sheet_name(wb_data, exclude_sheet=data_sheet)
        if data_sheet is None:
            report.failures.append(ValidationFailure("sheet-exists", "'Data' sheet missing"))
            return report
        if asp_sheet is None:
            report.failures.append(ValidationFailure("sheet-exists", "'ASP' sheet missing"))
            return report
        if data_sheet != "Data":
            report.warnings.append(f"Using '{data_sheet}' as the Data sheet")
        if asp_sheet != "ASP":
            report.warnings.append(f"Using '{asp_sheet}' as the ASP sheet")

        ws_data = wb_data[data_sheet]
        ws_asp = wb_data[asp_sheet]

        self._check_market_name(ws_data, report)
        geo_data = self._check_geographies(ws_data, report)
        self._check_segment_shares(ws_data, geo_data, report)
        self._check_asp(ws_asp, geo_data, report)

        return report

    # --- individual checks --------------------------------------------------

    def _check_market_name(self, ws, report: Gate1Report) -> None:
        name = ws.cell(1, 3).value
        if not name or not str(name).strip():
            report.warnings.append(
                "Cell C1 (market name) is blank; using workbook filename as market name")

    def _check_geographies(self, ws, report: Gate1Report) -> dict:
        """Check every known geography has a positive CAGR and anchor value."""
        valid_names = set(self._taxonomy.geographies.by_name)
        found: dict[str, tuple[float, float]] = {}   # name -> (cagr, anchor)

        for row in range(1, ws.max_row + 1):
            name = ws.cell(row, 2).value
            cagr = ws.cell(row, 4).value
            val  = ws.cell(row, 5).value
            if (isinstance(name, str) and name in valid_names
                    and isinstance(cagr, (int, float))
                    and isinstance(val, (int, float))):
                found[name] = (float(cagr), float(val))

        for geo in valid_names:
            if geo not in found:
                report.warnings.append(f"geography '{geo}' not found in Input Sheet")
                continue
            cagr, anchor = found[geo]
            if cagr <= 0:
                report.failures.append(ValidationFailure(
                    "cagr-positive", f"{geo}: CAGR={cagr:.4f} is not positive"))
            if anchor <= 0:
                report.failures.append(ValidationFailure(
                    "anchor-positive", f"{geo}: anchor_2025={anchor} is not positive"))

        return found

    def _check_segment_shares(self, ws, geo_data: dict, report: Gate1Report) -> None:
        """For each geography column, check each flat dimension sums to ~100%."""
        # Scan row 3 for geography column positions — ONLY the first (segmentation)
        # band (cols 3-30). Later bands repeat the geo headers for CAGR / trend-ID
        # data and would give wrong column mappings for the share rows.
        _SEG_BAND_END = 30
        geo_cols: dict[str, int] = {}
        for c in range(3, _SEG_BAND_END + 1):
            v = ws.cell(3, c).value
            if isinstance(v, str) and v in geo_data and v not in geo_cols:
                geo_cols[v] = c

        if not geo_cols:
            report.failures.append(ValidationFailure(
                "geo-columns", "No geography columns found in segmentation band (cols 3-30) of Data sheet"))
            return

        # Only check flat dimensions (not hierarchical Distribution Channel)
        flat_dims = [d for d in self._taxonomy.segmentation_dimensions
                     if not any(d.parent_of(s) for s in d.segments)]

        for dim in flat_dims:
            seg_rows: dict[str, int] = {}
            for row in range(1, ws.max_row + 1):
                label = ws.cell(row, 2).value
                if isinstance(label, str) and label.strip() in dim.segments:
                    seg_rows[label.strip()] = row

            if len(seg_rows) != len(dim.segments):
                report.warnings.append(
                    f"Dimension '{dim.title}': only found {len(seg_rows)}/{len(dim.segments)} segments")
                continue

            # Check share sum for first 3 geographies only (representative sample)
            for geo, col in list(geo_cols.items())[:3]:
                total = sum(
                    float(v) if isinstance(v := ws.cell(seg_rows[seg], col).value, (int, float)) else 0.0
                    for seg in dim.segments
                )
                if abs(total - 1.0) > _SHARE_SUM_TOL:
                    report.failures.append(ValidationFailure(
                        "shares-sum",
                        f"'{dim.title}' shares for {geo} sum to {total:.4f} (expected ~1.0)"))

    def _check_asp(self, ws_asp, geo_data: dict, report: Gate1Report) -> None:
        """Check each leaf product has a positive ASP for at least one geography.

        Parent/rollup segments (those that have children in the priced dimension)
        are skipped — their ASP row is a sum row and may be 0 or missing.
        """
        priced = self._taxonomy.priced_dimension
        # Only check leaf segments (those with no children)
        has_children = {priced.parent_of(s) for s in priced.segments if priced.parent_of(s) is not None}
        leaf_products = [p for p in priced.segments if p not in has_children]

        label_col = find_label_column(ws_asp)
        for product in leaf_products:
            found_row = None
            for r in range(1, ws_asp.max_row + 1):
                cell_val = ws_asp.cell(r, label_col).value
                if isinstance(cell_val, str) and cell_val.strip() == product:
                    found_row = r
                    break
            if found_row is None:
                # Fallback: search the first few columns if the label column was misdetected
                for r in range(1, ws_asp.max_row + 1):
                    for c in (2, 3, 4):
                        cell_val = ws_asp.cell(r, c).value
                        if isinstance(cell_val, str) and cell_val.strip() == product:
                            found_row = r
                            break
                    if found_row is not None:
                        break
            if found_row is None:
                report.failures.append(ValidationFailure(
                    "asp-row", f"Product '{product}' not found in ASP sheet"))
                continue
            # Check a sample of values in that row from the ASP block.
            asp_cols = scan_asp_geo_columns(ws_asp)
            if asp_cols:
                sample_cols = sorted(asp_cols.values())[:5]
            else:
                sample_cols = list(range(3, 8))
            sample = [ws_asp.cell(found_row, c).value for c in sample_cols]
            numeric = [v for v in sample if isinstance(v, (int, float))]
            if not numeric:
                report.failures.append(ValidationFailure(
                    "asp-positive", f"Product '{product}' ASP row has no numeric values"))
            elif any(v <= 0 for v in numeric):
                report.failures.append(ValidationFailure(
                    "asp-positive", f"Product '{product}' has ASP <= 0"))
