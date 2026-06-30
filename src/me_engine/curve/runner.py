"""Drive the Curve Agent across all geographies and score it against ground truth.

For each geography: ask the agent for a shape (with reasoning), build the value
path from the Input CAGR + anchor, and — when an ME file is available — score the
result against the human curve. Produces both the curves and a reasoning report.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from ..domain.runtime_taxonomy import RuntimeTaxonomy, default_taxonomy
from ..domain.series import Series
from ..domain.taxonomy import YEARS
from ..io.input_reader import GeoInput, InputSheetReader
from .agent import CurveAgent, CurveDecision, MarketContext
from .score import CurveScore, CurveScorer
from .shape import CurvePathBuilder


@dataclass(frozen=True, slots=True)
class CurveOutcome:
    geo: GeoInput
    decision: CurveDecision
    path: Series
    score: CurveScore | None

    def reasoning_line(self) -> str:
        tag = "FALLBACK" if self.decision.used_fallback else "AGENT"
        sc = f" | {self.score.summary()}" if self.score else ""
        return (f"[{tag}] {self.geo.name:16s} archetype={self.decision.archetype:24s} "
                f"peak={self.decision.peak_year} conf={self.decision.confidence:.2f}{sc}\n"
                f"        why: {self.decision.reasoning}")


class CurveRunner:
    """Orchestrates Input Sheet -> agent -> path -> score for every geography."""

    def __init__(self, agent: CurveAgent | None = None) -> None:
        self._agent = agent or CurveAgent()
        self._builder = CurvePathBuilder()
        self._scorer = CurveScorer()

    def run(self, input_path: Path | str,
            truth_paths: dict[str, Series] | None = None,
            taxonomy: RuntimeTaxonomy | None = None) -> list[CurveOutcome]:
        reader = InputSheetReader(input_path, taxonomy=taxonomy)
        inputs = reader.read()
        return [self._one(geo, truth_paths) for geo in inputs.values()]

    def _one(self, geo: GeoInput,
             truth_paths: dict[str, Series] | None) -> CurveOutcome:
        ctx = MarketContext(market_name="", geography=geo.name,
                            forecast_cagr=geo.forecast_cagr, region=geo.region)
        decision = self._agent.decide(ctx)
        path = self._builder.build(geo.anchor_2025, geo.forecast_cagr, decision.shape)
        score = None
        if truth_paths and geo.name in truth_paths:
            score = self._scorer.score(geo.name, decision.shape, path,
                                       truth_paths[geo.name])
        return CurveOutcome(geo=geo, decision=decision, path=path, score=score)


def truth_paths_from_me(me_path: Path | str) -> dict[str, Series]:
    """Extract each geography's total-value path from an existing ME workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(Path(me_path), data_only=True, read_only=True)
    paths: dict[str, Series] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = [ws.cell(6, c).value for c in range(4, 17)]
        if all(isinstance(v, (int, float)) for v in cells):
            paths[name] = Series(dict(zip(YEARS, (float(v) for v in cells))))
    return paths
