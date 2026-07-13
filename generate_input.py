"""Generate a new Input Sheet workbook and optionally run the ME pipeline.

This script creates a minimal valid Input Sheet workbook for a market using a
schema JSON or an existing template input workbook. It can also invoke
run_pipeline.py to generate an ME workbook from the generated input.
"""
from __future__ import annotations

import argparse
import json
import math
import random
import re
import subprocess
import sys
from pathlib import Path

# Ensure imports from src are available when running from repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent / "src"))

from openpyxl import Workbook, load_workbook
from me_engine.curve.generation_agent import GenerationAgent
from me_engine.domain.schema_loader import DimensionSchema, MarketSchema, load_schema


def slugify(name: str) -> str:
    normalized = name.strip().lower()
    for ch in [" ", "-", "/", "&", ".", ","]:
        normalized = normalized.replace(ch, "_")
    normalized = "_".join(part for part in normalized.split("_") if part)
    return normalized


def infer_schema_path_from_market(market_name: str, schemas_dir: Path) -> Path | None:
    slug = slugify(market_name)
    candidate = schemas_dir / f"{slug}.json"
    if candidate.exists():
        return candidate
    # Try best match by market_name inside JSON files.
    for path in sorted(schemas_dir.glob("*.json")):
        try:
            with path.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
            if data.get("market_name", "").strip().lower() == market_name.strip().lower():
                return path
        except Exception:
            continue
    return None


def infer_template_input_from_market(market_name: str, input_dir: Path) -> Path | None:
    if not input_dir.exists():
        return None
    normalized = slugify(market_name)
    for path in sorted(input_dir.glob("*.xlsx")):
        stem = slugify(path.stem)
        if normalized in stem or stem in normalized:
            return path
    return None


def load_schema_from_path(schema_path: Path | None, market_name: str | None) -> MarketSchema | None:
    if schema_path is None:
        return None
    data = json.loads(schema_path.read_text(encoding="utf-8"))
    if "dimensions" in data and "priced_dimension" in data:
        return load_schema(schema_path)

    # Legacy schema format: product_type + geographies
    dimensions: list[DimensionSchema] = []
    if "product_type" in data:
        dimensions.append(DimensionSchema(title="By Product Type", segments=tuple(data["product_type"])))
    if "geographies" not in data:
        raise ValueError("Schema must include geographies")
    return MarketSchema(
        market_name=data.get("market_name", market_name or ""),
        dimensions=tuple(dimensions),
        geographies=tuple(data["geographies"]),
        priced_dimension=dimensions[0].title if dimensions else "By Product Type",
        geography_tree=None,
    )


def load_schema_from_template(template_path: Path) -> MarketSchema:
    from me_engine.tools.extract_schema import extract_schema

    return extract_schema(template_path)


def equal_shares(count: int) -> list[float]:
    if count <= 0:
        return []
    share = 1.0 / count
    return [share] * count


def _segment_weight(label: str, index: int, geography_index: int) -> float:
    weight = 1.0 + 0.12 * index
    lowered = label.lower()
    if "premium" in lowered:
        weight *= 1.18
    if "value" in lowered:
        weight *= 0.9
    if "protein" in lowered:
        weight *= 1.06
    if "whey" in lowered:
        weight *= 1.12
    if "plant" in lowered:
        weight *= 0.96
    if "collagen" in lowered:
        weight *= 0.9
    return weight * (1.0 + 0.08 * ((geography_index + index) % 4))


def _normalize_weights(values: list[float]) -> list[float]:
    total = sum(values)
    if total <= 0:
        return [1.0 / len(values) if values else 0.0 for _ in values]

    normalized = [value / total for value in values]
    rounded = [round(value, 4) for value in normalized]
    if not rounded:
        return []

    diff = round(1.0 - sum(rounded), 4)
    rounded[-1] = round(rounded[-1] + diff, 4)
    return rounded


def _build_sibling_weights(labels: list[str], geographies: list[str]) -> list[list[float]]:
    weights: list[list[float]] = []
    for index, label in enumerate(labels):
        weights.append([_segment_weight(label, index, geo_index) for geo_index in range(len(geographies))])

    for geo_index in range(len(geographies)):
        geo_values = [weight[geo_index] for weight in weights]
        normalized = _normalize_weights(geo_values)
        for index, weight in enumerate(weights):
            weight[geo_index] = normalized[index]

    return weights


def build_segment_rows(dim: DimensionSchema, geographies: list[str]) -> list[tuple[str, list[float]]]:
    rows: list[tuple[str, list[float]]] = []
    if not any(parent is not None for parent in dim.parent.values()):
        sibling_weights = _build_sibling_weights(list(dim.segments), geographies)
        for seg, weights in zip(dim.segments, sibling_weights):
            rows.append((seg, [value for value in weights]))
        return rows

    parent_to_children: dict[str | None, list[str]] = {}
    for seg in dim.segments:
        parent = dim.parent.get(seg)
        parent_to_children.setdefault(parent, []).append(seg)

    top_level = parent_to_children.get(None, [])
    for parent in top_level:
        children = parent_to_children.get(parent, [])
        if not children:
            rows.append((parent, [1.0] * len(geographies)))
            continue

        child_weights = _build_sibling_weights(children, geographies)
        rows.append((parent, [1.0] * len(geographies)))
        for child, weights in zip(children, child_weights):
            rows.append((child, [value for value in weights]))

        grandchildren = [child for child in children if child in dim.parent.values()]
        for child in grandchildren:
            grand_children = [seg for seg, parent in dim.parent.items() if parent == child]
            if not grand_children:
                continue
            grand_weights = _build_sibling_weights(grand_children, geographies)
            for grand_child, weights in zip(grand_children, grand_weights):
                rows.append((grand_child, [value for value in weights]))

    for seg in dim.segments:
        if seg in top_level:
            continue
        if dim.parent.get(seg) is None and top_level:
            weights = [1.0] * len(geographies)
            rows.append((seg, weights))
            continue

    return rows


def build_geography_summary_rows(schema: MarketSchema, anchor: float) -> list[tuple[str, float, float, float]]:
    if schema.geography_tree and schema.geography_tree.children:
        leaves: list[tuple[str, float]] = []

        def collect(node, depth: int) -> None:
            if not node.children:
                leaves.append((node.name, 1.0 + 0.08 * depth))
                return
            for child in node.children:
                collect(child, depth + 1)

        collect(schema.geography_tree, 0)
        total = sum(weight for _, weight in leaves)
        normalized = [(name, weight / total) for name, weight in leaves] if total > 0 else [(name, 1.0 / len(leaves)) for name, _ in leaves]
        weights = {name: share for name, share in normalized}

        rows: list[tuple[str, float, float, float]] = []

        def walk(node, depth: int) -> None:
            children = list(node.children)
            if not children:
                share = weights.get(node.name, 0.0)
                cagr = round(0.045 + 0.003 * depth, 4)
                rows.append((node.name, share, cagr, round(anchor * share, 2)))
                return
            child_rows: list[tuple[str, float, float, float]] = []
            for child in children:
                walk(child, depth + 1)
            region_share = 0.0
            for child_name, child_share, _, _ in rows[-len(children):]:
                region_share += child_share
            cagr = round(0.042 + 0.002 * depth, 4)
            rows.append((node.name, region_share, cagr, round(anchor * region_share, 2)))

        for child in schema.geography_tree.children:
            walk(child, 1)

        # Keep rows in a stable order: region row first, then leaf rows.
        ordered: list[tuple[str, float, float, float]] = []
        seen: set[str] = set()
        for row in rows:
            if row[0] in seen:
                continue
            ordered.append(row)
            seen.add(row[0])
        for row in rows:
            if row[0] not in seen:
                ordered.append(row)
        return ordered

    geographies = list(schema.geographies)
    if not geographies:
        return []
    region_names = ["North America", "Europe", "Asia Pacific", "Latin America"]
    rows: list[tuple[str, float, float, float]] = []
    for idx, geo in enumerate(geographies):
        region = region_names[idx % len(region_names)]
        share = 0.18 + 0.04 * idx
        if share > 0.95:
            share = 0.95
        rows.append((geo, round(share, 4), round(0.05 + 0.002 * idx, 4), round(anchor * share, 2)))
    if len(geographies) >= 2:
        rows.insert(0, (region_names[0], round(sum(item[1] for item in rows[:2]), 4), 0.057, round(anchor * sum(item[1] for item in rows[:2]), 2)))
    return rows


def build_cagr_values(geographies: list[str], base: float, variation: float) -> list[float]:
    return [round(base + variation * math.sin(i * 0.7), 4) for i in range(len(geographies))]


def build_anchor_values(geographies: list[str], base: float) -> list[float]:
    return [round(base * (1.0 + 0.05 * (i % 5)), 2) for i in range(len(geographies))]


def build_asp_rows(priced_dim: str, products: list[str], geographies: list[str]) -> list[tuple[str, list[float]]]:
    rows: list[tuple[str, list[float]]] = []
    for idx, product in enumerate(products, start=1):
        base = 10.0 + idx * 2.0
        row = [round(base + 0.5 * j, 2) for j in range(len(geographies))]
        rows.append((product, row))
    return rows


def create_input_workbook(schema: MarketSchema, market_name: str, output_path: Path, cagr: float, anchor: float) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.cell(1, 2).value = "Market Name"
    ws.cell(1, 3).value = market_name

    geographies = list(schema.geographies)
    for col_index, geo in enumerate(geographies, start=3):
        ws.cell(3, col_index).value = geo
    for col_index, geo in enumerate(geographies, start=3 + len(geographies) + 1):
        ws.cell(3, col_index).value = geo
    for col_index, geo in enumerate(geographies, start=3 + 2 * (len(geographies) + 1)):
        ws.cell(3, col_index).value = geo

    current_row = 4
    for dim in schema.dimensions:
        ws.cell(current_row, 2).value = dim.title
        current_row += 1
        for label, values in build_segment_rows(dim, geographies):
            ws.cell(current_row, 2).value = label
            for col_index, value in enumerate(values, start=3):
                ws.cell(current_row, col_index).value = round(value, 4)
            current_row += 1
        current_row += 1

    ws.cell(current_row, 2).value = "Market Size US$ Mn in 2025"
    ws.cell(current_row, 3).value = "Share"
    ws.cell(current_row, 4).value = "CAGR"
    ws.cell(current_row, 5).value = "Value"
    current_row += 1

    for label, share, cagr_value, value in build_geography_summary_rows(schema, anchor):
        ws.cell(current_row, 2).value = label
        ws.cell(current_row, 3).value = round(share, 4)
        ws.cell(current_row, 4).value = round(cagr_value, 4)
        ws.cell(current_row, 5).value = round(value, 2)
        current_row += 1

    ws.cell(current_row, 2).value = ""
    current_row += 1
    ws.cell(current_row, 2).value = "Share"
    ws.cell(current_row, 3).value = "CAGR"
    ws.cell(current_row, 4).value = "Value"
    current_row += 1

    for label, share, cagr_value, value in build_geography_summary_rows(schema, anchor):
        if label in geographies:
            ws.cell(current_row, 2).value = label
            ws.cell(current_row, 3).value = round(share, 4)
            ws.cell(current_row, 4).value = round(cagr_value, 4)
            ws.cell(current_row, 5).value = round(value, 2)
            current_row += 1

    asp = wb.create_sheet("ASP")
    asp.cell(1, 2).value = "Market Name"
    asp.cell(1, 3).value = market_name
    asp.cell(3, 2).value = f"By {schema.priced_dimension}"
    asp.cell(4, 2).value = "Product"
    for col_index, geo in enumerate(geographies, start=3):
        asp.cell(4, col_index).value = geo

    priced_dim = next((dim for dim in schema.dimensions if dim.title == schema.priced_dimension), schema.dimensions[0] if schema.dimensions else None)
    product_segments = list(priced_dim.segments) if priced_dim is not None else []
    for row_index, (product, values) in enumerate(build_asp_rows(schema.priced_dimension, product_segments, geographies), start=5):
        asp.cell(row_index, 2).value = product
        for col_index, value in enumerate(values, start=3):
            asp.cell(row_index, col_index).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def run_pipeline(input_path: Path, output_path: Path, reference_path: Path | None, report_path: Path | None, save_metadata: bool) -> int:
    args = [sys.executable, str(Path(__file__).resolve().parent / "run_pipeline.py"), "--input", str(input_path), "--output", str(output_path)]
    if reference_path is not None:
        args.extend(["--reference", str(reference_path)])
    if report_path is not None:
        args.extend(["--save-report", str(report_path)])
    if save_metadata:
        args.append("--save-metadata")

    proc = subprocess.run(args, capture_output=True, text=True)
    print(proc.stdout)
    print(proc.stderr, file=sys.stderr)
    return proc.returncode


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate an Input Sheet workbook from market metadata.")
    p.add_argument("--market-name", required=True, help="Market name for the generated input sheet")
    p.add_argument("--schema", default=None, help="Optional schema JSON file to use for market structure")
    p.add_argument("--template-input", default=None,
                   help="Optional existing Input Sheet workbook to infer schema from")
    p.add_argument("--output", default=None,
                   help="Output path for the generated Input Sheet workbook")
    p.add_argument("--output-folder", default="agentGeneratedInputSheet",
                   help="Folder to write generated input sheets when --output is not provided")
    p.add_argument("--cagr", type=float, default=0.08,
                   help="Base CAGR %% to write into the generated input sheet")
    p.add_argument("--anchor", type=float, default=100.0,
                   help="Base 2025 anchor value for each geography")
    p.add_argument("--run-pipeline", action="store_true",
                   help="Run run_pipeline.py immediately after generating the input sheet")
    p.add_argument("--reference", default=None,
                   help="Optional reference ME workbook for pipeline comparison")
    p.add_argument("--report", default=None,
                   help="Optional JSON report path for the pipeline run")
    p.add_argument("--save-metadata", action="store_true",
                   help="Pass --save-metadata when running the pipeline")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    market_name = args.market_name.strip()

    schema_path = Path(args.schema) if args.schema else None
    template_input = Path(args.template_input) if args.template_input else None
    if schema_path is None:
        default_schema = infer_schema_path_from_market(market_name, Path("schemas"))
        if default_schema is not None:
            schema_path = default_schema

    if template_input is None:
        candidate = infer_template_input_from_market(market_name, Path("my_inputs"))
        if candidate is not None:
            template_input = candidate

    schema = None
    if schema_path is not None:
        schema = load_schema_from_path(schema_path, market_name)
    elif template_input is not None:
        schema = load_schema_from_template(template_input)

    if schema is None:
        agent = GenerationAgent(root=Path(__file__).resolve().parent)
        schema = agent._load_schema(None, market_name)

    if args.output:
        output_path = Path(args.output)
        if output_path.exists() and output_path.is_dir():
            output_path = output_path / f"Data File - {market_name}.xlsx"
    else:
        output_path = Path(args.output_folder) / f"Data File - {market_name}.xlsx"
    generated_path = create_input_workbook(schema, market_name, output_path, args.cagr, args.anchor)
    print(f"Generated Input Sheet: {generated_path}")

    if args.run_pipeline:
        output_me = Path(f"my_outputs/ME-{market_name}.xlsx")
        report_path = Path(args.report) if args.report else Path(f"output/report-{market_name}.json")
        reference_path = Path(args.reference) if args.reference else None
        rc = run_pipeline(generated_path, output_me, reference_path, report_path, args.save_metadata)
        if rc != 0:
            print(f"Pipeline failed with exit code {rc}")
            return rc

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
