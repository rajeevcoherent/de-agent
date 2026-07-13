from pathlib import Path
import importlib
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from me_engine.curve.generation_agent import GenerationAgent
from me_engine.domain.schema_loader import DimensionSchema, GeographyNodeSchema, MarketSchema
from generate_input import create_input_workbook


class _FakeLLM:
    def complete_json(self, system: str, user: str) -> dict:
        return {
            "market_name": "Protein Supplements Market",
            "dimensions": [
                {
                    "title": "By Product Type",
                    "segments": ["Protein Supplements", "Whey Protein", "Plant Protein", "Collagen Protein"],
                    "parent": {
                        "Whey Protein": "Protein Supplements",
                        "Plant Protein": "Protein Supplements",
                        "Collagen Protein": "Protein Supplements",
                    },
                },
                {
                    "title": "By Consumer",
                    "segments": ["Consumers", "Athletes", "General Wellness", "Seniors"],
                    "parent": {
                        "Athletes": "Consumers",
                        "General Wellness": "Consumers",
                        "Seniors": "Consumers",
                    },
                },
            ],
            "geographies": ["U.S.", "Canada", "Germany", "U.K."],
            "priced_dimension": "By Product Type",
            "geography_tree": {
                "name": "World",
                "children": [
                    {"name": "North America", "children": [{"name": "U.S."}, {"name": "Canada"}]},
                    {"name": "Europe", "children": [{"name": "Germany"}, {"name": "U.K."}]},
                ],
            },
        }


class _FakeEvidence:
    def search(self, query: str) -> list[str]:
        return ["protein supplements market segmentation and regional demand"]


def test_generation_agent_returns_blueprint() -> None:
    agent = GenerationAgent()
    plan = agent.plan("Global Adhesives & Sealants Market")

    assert plan.market_name == "Global Adhesives & Sealants Market"
    assert plan.output_name.endswith(".xlsx")
    assert plan.output_name.startswith("Data File -")
    assert plan.output_path.endswith(".xlsx")


def test_create_input_workbook_emits_rich_health_market_structure(tmp_path: Path) -> None:
    schema = MarketSchema(
        market_name="Protein Supplements Market",
        dimensions=(
            DimensionSchema(title="By Product Type", segments=("Whey Protein", "Plant Protein", "Collagen Protein")),
            DimensionSchema(title="By Consumer", segments=("Athletes", "General Wellness", "Seniors")),
        ),
        geographies=("U.S.", "Canada", "Germany", "U.K."),
        priced_dimension="By Product Type",
        geography_tree=None,
    )
    output_path = tmp_path / "protein_input.xlsx"

    create_input_workbook(schema, schema.market_name, output_path, 0.08, 100.0)

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Data"]

    assert ws.cell(1, 3).value == schema.market_name
    assert ws.cell(3, 3).value == "U.S."
    assert any(ws.cell(row, 2).value == "By Product Type" for row in range(1, 20))
    assert any(ws.cell(row, 2).value in {"North America", "Western Europe", "Europe"} for row in range(1, 80))
    assert any(ws.cell(row, 2).value == "U.S." for row in range(1, 80))


def test_generation_agent_uses_llm_plan_to_build_schema(tmp_path: Path) -> None:
    agent = GenerationAgent(
        root=Path(__file__).resolve().parents[1],
        llm=_FakeLLM(),
        evidence=_FakeEvidence(),
    )
    agent._outputs_dir = tmp_path / "agentGeneratedInputSheet"

    generated_path = agent.materialize("Protein Supplements Market")

    assert generated_path.exists()
    wb = load_workbook(generated_path, data_only=True)
    ws = wb["Data"]
    assert ws.cell(1, 3).value == "Protein Supplements Market"
    assert any(ws.cell(row, 2).value == "By Product Type" for row in range(1, 20))
    assert any(ws.cell(row, 2).value == "Protein Supplements" for row in range(1, 80))
    assert any(ws.cell(row, 2).value == "Whey Protein" for row in range(1, 80))
    assert any(ws.cell(row, 2).value == "U.S." for row in range(1, 80))
    assert any(ws.cell(row, 2).value == "North America" for row in range(1, 120))


def test_create_input_workbook_uses_geography_tree_and_hierarchy(tmp_path: Path) -> None:
    schema = MarketSchema(
        market_name="Protein Supplements Market",
        dimensions=(
            DimensionSchema(
                title="By Product Type",
                segments=("Protein", "Whey Protein", "Plant Protein"),
                parent={"Whey Protein": "Protein", "Plant Protein": "Protein"},
            ),
        ),
        geographies=("U.S.", "Canada", "Germany", "U.K."),
        priced_dimension="By Product Type",
        geography_tree=GeographyNodeSchema(
            name="World",
            children=(
                GeographyNodeSchema(
                    name="North America",
                    children=(
                        GeographyNodeSchema(name="U.S."),
                        GeographyNodeSchema(name="Canada"),
                    ),
                ),
                GeographyNodeSchema(
                    name="Western Europe",
                    children=(
                        GeographyNodeSchema(name="Germany"),
                        GeographyNodeSchema(name="U.K."),
                    ),
                ),
            ),
        ),
    )
    output_path = tmp_path / "hierarchy_input.xlsx"

    create_input_workbook(schema, schema.market_name, output_path, 0.08, 100.0)

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Data"]

    assert any(ws.cell(row, 2).value == "Protein" for row in range(1, 50))
    assert any(ws.cell(row, 2).value == "Whey Protein" for row in range(1, 50))
    assert any(ws.cell(row, 2).value == "Western Europe" for row in range(1, 120))
    assert any(ws.cell(row, 2).value == "U.S." for row in range(1, 120))


def test_flat_dimension_rows_sum_to_one_for_each_geography(tmp_path: Path) -> None:
    schema = MarketSchema(
        market_name="Mobile Device Market",
        dimensions=(
            DimensionSchema(title="By Channel", segments=("Direct", "Retail", "Online")),
        ),
        geographies=("U.S.", "Germany", "Japan", "China"),
        priced_dimension="By Channel",
        geography_tree=None,
    )
    output_path = tmp_path / "flat_dimension_input.xlsx"

    create_input_workbook(schema, schema.market_name, output_path, 0.08, 100.0)

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Data"]
    row_lookup = {ws.cell(row, 2).value: row for row in range(1, ws.max_row + 1) if isinstance(ws.cell(row, 2).value, str)}

    for col in range(3, 7):
        total = sum(float(ws.cell(row_lookup[segment], col).value) for segment in ("Direct", "Retail", "Online"))
        assert total == pytest.approx(1.0, abs=1e-9)


def test_config_import_is_silent_without_api_keys(capsys: pytest.CaptureFixture[str]) -> None:
    import me_engine.curve.config as config_module

    reload_module = importlib.reload(config_module)
    assert reload_module is config_module

    captured = capsys.readouterr()
    assert captured.out == ""
    assert captured.err == ""
