"""Step-by-step debug — find where the pipeline hangs."""
import sys, time
sys.path.insert(0, 'src')
from pathlib import Path

INPUT  = "my_inputs/Data file-Global High Energy Laser Market.xlsx"
OUTPUT = "my_outputs/debug_laser.xlsx"
REF    = "my_reference/ME - Global High Energy Laser Market.xlsx"

print("Step 1: build taxonomy...", flush=True)
t = time.time()
from me_engine.domain.runtime_taxonomy import build_taxonomy_from_workbook
taxonomy = build_taxonomy_from_workbook(INPUT)
print(f"  OK ({time.time()-t:.1f}s) geos={len(taxonomy.geographies.by_name)} dims={len(taxonomy.segmentation_dimensions)}", flush=True)

print("Step 2: build drivers (no-agents)...", flush=True)
t = time.time()
from me_engine.io.input_drivers import InputDriverBuilder
builder = InputDriverBuilder(INPUT, truth_me=None, use_segmentation_agent=False, taxonomy=taxonomy)
drivers = builder.build()
print(f"  OK ({time.time()-t:.1f}s) geos={len(drivers.geographies)}", flush=True)

print("Step 3: assemble...", flush=True)
t = time.time()
from me_engine.assembly.assembler import Assembler
result = Assembler(taxonomy).assemble(drivers)
print(f"  OK ({time.time()-t:.1f}s) geos={len(result.geographies)}", flush=True)

print("Step 4: check reference sheetnames...", flush=True)
from openpyxl import load_workbook
ref_wb = load_workbook(REF, read_only=True)
print(f"  Reference sheets: {ref_wb.sheetnames[:5]}...", flush=True)
result_geos = list(result.geographies.keys())
print(f"  Result geos (first 5): {result_geos[:5]}", flush=True)
overlap = [g for g in result_geos if g in ref_wb.sheetnames]
print(f"  Overlap with reference: {len(overlap)}", flush=True)

print("Step 5: write output...", flush=True)
t = time.time()
Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)
from me_engine.io.me_writer import MEWorkbookWriter
writer = MEWorkbookWriter(REF)
writer.write(result, OUTPUT)
print(f"  OK ({time.time()-t:.1f}s)", flush=True)

print("Step 6: Gate 3 diff...", flush=True)
t = time.time()
from me_engine.validation.gate3 import OutputValidator
gate3 = OutputValidator(taxonomy).validate(result, REF, {})
print(f"  OK ({time.time()-t:.1f}s)", flush=True)
if gate3.diff:
    print(f"  Mean error: {gate3.diff.mean_rel_error:.2%}")
    print(f"  Within 5%:  {gate3.diff.cells_within_5pct:.1%}")

print("\nDONE", flush=True)
